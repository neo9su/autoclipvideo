import asyncio
import hashlib
import json
import logging
import os
import re
import shutil
from contextlib import asynccontextmanager
from pathlib import Path

# Load .env file if present (before any module reads os.environ)
_env_path = os.path.join(os.path.dirname(__file__), ".env")
if os.path.exists(_env_path):
    with open(_env_path, encoding='utf-8') as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _v = _line.split("=", 1)
                os.environ.setdefault(_k.strip(), _v.strip())
from typing import Optional, Set

import aiosqlite
import httpx
from datetime import datetime, timedelta

from fastapi import Body, FastAPI, File, Form, HTTPException, Query, Request, UploadFile, WebSocket, WebSocketDisconnect
from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

from db import init_db, DB_PATH, aio_connect
from models import RoomCreate, Room, Recording, ProductCreate, ProductUpdate, PublishAccountCreate, PublishTaskCreate, BatchScheduleCreate
from monitor import MonitorManager
from transcribe import poll_transcriptions, _run_editor, _clip_progress, get_clip_queue, update_job_priority, cancel_clip_job, pause_clip_job, resume_clip_job, _job_submit_times, _job_durations, _poll_state, flush_poll, POLL_INTERVAL, RECORDINGS_DIR, backfill_auto_merge
from analyzer import merge_group
from sync import sync_file
from gpu_execution import require_remote_gpu
from thumbnail import generate_thumbnail
from meta_generator import generate_meta, match_product
from publish_scheduler import poll_publish_tasks
from video_path_resolver import resolve_artifact_path
from srt_resolver import resolve_srt_path

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
logger = logging.getLogger(__name__)

# ── Deployment role (must be set BEFORE lifespan reads it) ──────────────────
# control-plane = Mac light instance (port 8099): API/static files only, no media pipelines
# gpu-backend   = GPU server (port 8899): full media pipeline with remote GPU access
_DEPLOYMENT_ROLE = os.environ.get("DEPLOYMENT_ROLE", "gpu-backend")
if _DEPLOYMENT_ROLE not in {"gpu-backend", "control-plane"}:
    raise RuntimeError(
        "DEPLOYMENT_ROLE must be 'gpu-backend' or 'control-plane'"
    )
IS_CONTROL_PLANE = _DEPLOYMENT_ROLE == "control-plane"

# Remote GPU backend URL — canonical entry for media APIs (qianchuan, director, etc.)
REMOTE_BACKEND_URL = os.environ.get("REMOTE_BACKEND_URL", "http://10.190.0.203:8899")

STALE_OPEN_RECORDING_HOURS = 6

# WebSocket connections
_ws_clients: Set[WebSocket] = set()

_API_CACHE_TTL_SECONDS = 5.0
_rooms_cache: tuple[float, list[dict]] | None = None
_groups_cache: tuple[float, list[dict]] | None = None


async def broadcast(message: dict):
    payload = json.dumps(message, ensure_ascii=False, separators=(",", ":"))
    dead = set()
    for ws in _ws_clients:
        try:
            await ws.send_text(payload)
        except Exception:
            dead.add(ws)
    _ws_clients.difference_update(dead)


monitor = MonitorManager(
    broadcast_fn=broadcast,
    allow_local_media=not IS_CONTROL_PLANE,
)


def _recording_file_path(filename: str | None) -> str | None:
    """Return the expected local MP4 path for a recording filename."""
    if not filename:
        return None
    return os.path.join(RECORDINGS_DIR, filename)


def _recording_file_exists(filename: str | None) -> bool:
    """Return True when the recording source file is present locally."""
    path = _recording_file_path(filename)
    return bool(path and os.path.exists(path))


def _parse_recording_time(value: str | None) -> datetime | None:
    """Parse DB timestamps stored with either ISO 'T' or SQLite space separators."""
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace(" ", "T"))
    except ValueError:
        return None


def _is_finished_unsynced_upload_candidate(row) -> bool:
    """Return True when an unsynced recording can actually enter upload processing.

    The transcription poll loop only processes finished recordings with a real duration.
    Mirroring that guard here prevents crash-left placeholders (end_time NULL and no MP4)
    from inflating the queue denominator and making the UI look stuck at 98%.
    """
    if row["transcribed"] != 0 or row["synced"] != 0 or row["local_deleted"] != 0:
        return False
    if not row["end_time"] or row["end_time"] == row["start_time"]:
        return False
    return _recording_file_exists(row["filename"])


async def _cleanup_stale_open_recording_placeholders(max_age_hours: int = STALE_OPEN_RECORDING_HOURS) -> int:
    """Mark old open recording placeholders without MP4 files as skipped.

    Active recordings intentionally have end_time NULL, so only records older than the
    grace window and missing their source file are cleaned up. This handles recorder
    crashes without removing the currently-recording segment.
    """
    cutoff = datetime.now() - timedelta(hours=max_age_hours)
    stale_ids: list[int] = []

    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            """SELECT id, filename, start_time
               FROM recordings
               WHERE transcribed = 0
                 AND synced = 0
                 AND local_deleted = 0
                 AND end_time IS NULL"""
        ) as cur:
            rows = await cur.fetchall()

    for rec in rows:
        started_at = _parse_recording_time(rec["start_time"])
        if started_at is None or started_at > cutoff:
            continue
        if _recording_file_exists(rec["filename"]):
            continue
        stale_ids.append(rec["id"])

    if not stale_ids:
        return 0

    placeholders = ",".join("?" * len(stale_ids))
    reason = "stale open recording placeholder: missing local MP4"
    async with aio_connect() as db:
        await db.execute(
            f"""UPDATE recordings
                SET local_deleted = 1,
                    transcribed = -1,
                    transcribe_error = ?,
                    skip_reason = ?
                WHERE id IN ({placeholders})""",
            [reason, reason, *stale_ids],
        )
        await db.commit()
    logger.warning(
        "Marked %d stale open recording placeholder(s) as skipped after %dh grace window",
        len(stale_ids),
        max_age_hours,
    )
    return len(stale_ids)


async def _reset_stuck_clip_tasks():
    """On startup, reset recordings stuck at clipped=1 (server killed mid-clip).
    If clip_filename is already set → mark done (clipped=2); otherwise reset to pending (clipped=0).
    Also verify clipped=2 recordings still have their files on disk (GPU server may have restarted
    and purged clips); if the file is missing, reset to clipped=0 so it gets re-queued.
    
    Also reset clip_groups stuck at director_status=1 or creative_status=1 (server killed mid-pipeline)."""
    recordings_dir = os.path.join(os.path.dirname(__file__), "..", "recordings")
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        r1 = await db.execute(
            "UPDATE recordings SET clipped=2 WHERE clipped=1 AND clip_filename IS NOT NULL"
        )
        r2 = await db.execute(
            "UPDATE recordings SET clipped=0 WHERE clipped=1 AND clip_filename IS NULL"
        )
        await db.commit()
        if r1.rowcount:
            logger.info(f"Reset {r1.rowcount} stuck clip task(s) to done (clip_filename present)")
        if r2.rowcount:
            logger.info(f"Reset {r2.rowcount} stuck clip task(s) to pending (no clip_filename)")

        # Verify clipped=2 files actually exist on disk.
        # Only reset if the source MP4 still exists (otherwise can't re-clip anyway).
        # Skip local_deleted=1 recordings (user intentionally cleared them).
        async with db.execute(
            """SELECT id, filename, clip_filename FROM recordings
               WHERE clipped = 2 AND clip_filename IS NOT NULL AND local_deleted = 0"""
        ) as cur:
            done_rows = await cur.fetchall()

        missing = [
            r["id"] for r in done_rows
            if not os.path.exists(os.path.join(recordings_dir, r["clip_filename"]))
            and os.path.exists(os.path.join(recordings_dir, r["filename"]))
        ]
        if missing:
            placeholders = ",".join("?" * len(missing))
            await db.execute(
                f"UPDATE recordings SET clipped=0, clip_filename=NULL WHERE id IN ({placeholders})",
                missing,
            )
            await db.commit()
            logger.warning(f"Reset {len(missing)} recording(s) with missing clip files back to pending (MP4 still present)")

        # Reset clip_groups stuck at merge_status=1 (server killed mid-merge).
        # Re-check: if any recording in the group is now clipped=2, leave it — the
        # merge may have actually completed; otherwise reset to 0 so it can retry.
        r3 = await db.execute(
            """UPDATE clip_groups SET merge_status = 0
               WHERE merge_status = 1 AND merged_filename IS NULL AND director_final_video IS NULL"""
        )
        await db.commit()
        if r3.rowcount:
            logger.info(f"Reset {r3.rowcount} stuck merge(s) to pending")

        # Reset clip_groups stuck at director_status=1, creative_status=1, or qianchuan_status=1 (server killed mid-pipeline)
        r4 = await db.execute(
            "UPDATE clip_groups SET director_status = 0 WHERE director_status = 1"
        )
        r5 = await db.execute(
            "UPDATE clip_groups SET creative_status = 0 WHERE creative_status = 1"
        )
        r6 = await db.execute(
            "UPDATE clip_groups SET qianchuan_status = 0 WHERE qianchuan_status = 1"
        )
        await db.commit()
        if r4.rowcount:
            logger.info(f"Reset {r4.rowcount} stuck director pipeline(s) to pending")
        if r5.rowcount:
            logger.info(f"Reset {r5.rowcount} stuck creative pipeline(s) to pending")
        if r6.rowcount:
            logger.info(f"Reset {r6.rowcount} stuck qianchuan pipeline(s) to pending")


async def _startup_trigger_pipelines():
    """On startup, trigger director+creative pipelines for groups that have
    classic done but pipelines not yet started.
    Strategy: run director first, then creative only after director is done.
    Throttled to batches of 10 to avoid flooding the GPU TTS/concattenation queue.
    
    If GPU service is offline, skip dispatching — the periodic dispatch will retry
    when GPU comes back online."""
    from gpu_state import is_online as gpu_is_online
    if not gpu_is_online():
        logger.info("Startup pipeline trigger: GPU service offline — skipping dispatch")
        return
    from transcribe import _run_director_pipeline, _run_creative_pipeline, _extract_srt_for_director
    BATCH_SIZE = 10
    try:
        async with aio_connect() as db:
            db.row_factory = aiosqlite.Row
            # Director groups: classic done, director not done
            # Exclude groups with unrecoverable errors (missing files, no recordings, etc.)
            async with db.execute(
                """SELECT id FROM clip_groups
                   WHERE classic_status = 2
                     AND (director_status IN (0, -1, -3))
                     AND (
                       director_error IS NULL OR director_error = '' OR (
                         director_error NOT LIKE '%no recordings%'
                         AND director_error NOT LIKE '%recording files missing%'
                         AND director_error NOT LIKE '%physically deleted%'
                         AND director_error NOT LIKE '%无录像文件%'
                         AND director_error NOT LIKE '%无合并素材%'
                         AND director_error NOT LIKE '%no SRT content%'
                         AND director_error NOT LIKE '%SRT%'
                         AND director_error NOT LIKE '%video_clips is EMPTY%'
                         AND director_error NOT LIKE '%时长%'
                       )
                     )
                   ORDER BY id DESC"""
            ) as cur:
                pending_director = [r["id"] for r in await cur.fetchall()]

            # Director mode requires SRT content. Old records often have DB status
            # transcribed=2 but no .srt file on disk, so mark those as permanent
            # skips before dispatching to avoid endless startup retries.
            srt_ready_director = []
            skipped_no_srt = 0
            for gid in pending_director:
                if await _extract_srt_for_director(gid):
                    srt_ready_director.append(gid)
                else:
                    await db.execute(
                        "UPDATE clip_groups SET director_status = -2, director_error = ? WHERE id = ?",
                        ("no SRT content available", gid),
                    )
                    skipped_no_srt += 1
            if skipped_no_srt:
                await db.commit()
                logger.warning(f"Startup pipeline trigger: skipped {skipped_no_srt} director groups with missing SRT")
            pending_director = srt_ready_director
            # Creative groups: classic done, director done, creative not done
            # Exclude groups with unrecoverable errors
            async with db.execute(
                """SELECT id FROM clip_groups
                   WHERE classic_status = 2
                     AND director_status = 2
                     AND (creative_status IN (0, -1, -3) OR creative_status IS NULL)
                     AND (creative_error IS NULL OR creative_error = '')
                     AND creative_error NOT LIKE '%duration 0%'
                     AND creative_error NOT LIKE '%no recordings%'
                   ORDER BY id DESC"""
            ) as cur:
                pending_creative = [r["id"] for r in await cur.fetchall()]
            # Qianchuan groups: classic done, qianchuan not done
            # Exclude groups with unrecoverable errors
            async with db.execute(
                """SELECT id FROM clip_groups
                   WHERE classic_status = 2
                     AND qianchuan_status IN (0, -1, -3, -4)
                     AND (qianchuan_error IS NULL OR qianchuan_error = '' OR (
                       qianchuan_error NOT LIKE '%no recordings%'
                       AND qianchuan_error NOT LIKE '%recording files missing%'
                       AND qianchuan_error NOT LIKE '%physically deleted%'
                       AND qianchuan_error NOT LIKE '%无录像文件%'
                       AND qianchuan_error NOT LIKE '%无素材%'
                       AND qianchuan_error NOT LIKE '%时长%'
                       AND qianchuan_error NOT LIKE '%video_clips is EMPTY%'
                     ))
                   ORDER BY id DESC"""
            ) as cur:
                pending_qianchuan = [r["id"] for r in await cur.fetchall()]
        if not pending_director and not pending_creative and not pending_qianchuan:
            logger.info("Startup pipeline trigger: no pending groups")
            return
        logger.info(
            "Startup pipeline trigger: launching %d director + %d creative + %d qianchuan groups",
            len(pending_director), len(pending_creative), len(pending_qianchuan),
        )
        # Phase 1: dispatch director groups first
        for i in range(0, len(pending_director), BATCH_SIZE):
            batch = pending_director[i:i + BATCH_SIZE]
            for gid in batch:
                asyncio.create_task(_run_director_pipeline(gid))
                await asyncio.sleep(0.1)
            logger.info("Startup pipeline trigger: dispatched director batch %d (%d groups)", i // BATCH_SIZE + 1, len(batch))
            if i + BATCH_SIZE < len(pending_director):
                await asyncio.sleep(1)
        # Phase 2: dispatch creative groups (director already done)
        for i in range(0, len(pending_creative), BATCH_SIZE):
            batch = pending_creative[i:i + BATCH_SIZE]
            for gid in batch:
                asyncio.create_task(_run_creative_pipeline(gid))
                await asyncio.sleep(0.1)
            logger.info("Startup pipeline trigger: dispatched creative batch %d (%d groups)", i // BATCH_SIZE + 1, len(batch))
            if i + BATCH_SIZE < len(pending_creative):
                await asyncio.sleep(1)
        # Phase 3: dispatch qianchuan groups (independent of director/creative)
        if pending_qianchuan:
            try:
                from api_v2 import _run_qianchuan_pipeline
            except ImportError as ie:
                logger.warning("Startup pipeline trigger: api_v2 not available — skipping qianchuan dispatch (error: %s)", ie)
            else:
                for i in range(0, len(pending_qianchuan), BATCH_SIZE):
                    batch = pending_qianchuan[i:i + BATCH_SIZE]
                    for gid in batch:
                        asyncio.create_task(_run_qianchuan_pipeline(gid))
                        await asyncio.sleep(0.1)
                    logger.info("Startup pipeline trigger: dispatched qianchuan batch %d (%d groups)", i // BATCH_SIZE + 1, len(batch))
                    if i + BATCH_SIZE < len(pending_qianchuan):
                        await asyncio.sleep(1)
    except Exception as e:
        logger.error(f"Startup pipeline trigger failed: {e}")


async def _memory_monitor(broadcast_fn=None):
    """Monitor system RAM every 30 s.

    When used memory exceeds MEM_WARN_GB:
      - Set memory_pressure=True to pause new clip dispatches
      - Call gc.collect() to release Python-held objects
      - Broadcast a warning to the frontend

    When memory recovers below MEM_RECOVER_GB the pressure flag is cleared.
    """
    import gc

    # macOS: vm.used includes file cache; vm.available is what actually matters
    # GPU jobs run on remote server — local memory cost is minimal (file upload/download only)
    MEM_AVAIL_MIN_GB = 0.4  # pause only when truly critical (< 400 MB free)
    MEM_AVAIL_OK_GB  = 0.8  # recover when > 800 MB available
    INTERVAL         = 10   # seconds

    try:
        import psutil
    except ImportError:
        logger.warning("psutil not installed – memory monitor disabled. Run: pip install psutil")
        return

    from transcribe import set_memory_pressure

    warned = False
    while True:
        try:
            vm = psutil.virtual_memory()
            avail_gb = vm.available / 1024 ** 3

            if avail_gb < MEM_AVAIL_MIN_GB:
                gc.collect()
                vm2 = psutil.virtual_memory()
                avail_after = vm2.available / 1024 ** 3
                set_memory_pressure(True)
                msg = (
                    f"[内存警告] 可用内存 {avail_gb:.1f}GB 低于 {MEM_AVAIL_MIN_GB}GB 阈值，"
                    f"已暂停新剪辑任务派发 (gc后 {avail_after:.1f}GB)"
                )
                logger.warning(msg)
                if broadcast_fn and not warned:
                    await broadcast_fn({
                        "type": "memory_warning",
                        "avail_gb": round(avail_gb, 1),
                        "limit_gb": MEM_AVAIL_MIN_GB,
                        "msg": msg,
                    })
                warned = True

            elif warned and avail_gb > MEM_AVAIL_OK_GB:
                set_memory_pressure(False)
                logger.info(f"[内存恢复] 可用内存 {avail_gb:.1f}GB > {MEM_AVAIL_OK_GB}GB，已恢复剪辑任务派发")
                if broadcast_fn:
                    await broadcast_fn({
                        "type": "memory_recovered",
                        "avail_gb": round(avail_gb, 1),
                    })
                warned = False

        except asyncio.CancelledError:
            return
        except Exception as e:
            logger.debug(f"Memory monitor error: {e}")

        await asyncio.sleep(INTERVAL)


async def _on_gpu_online():
    """
    Auto-retry recently-failed clip jobs when GPU comes back online.
    Targets recordings with clipped=-1, no skip_reason, SRT available, failed in last 24h.
    """
    try:
        from datetime import timedelta
        cutoff = (datetime.utcnow() - timedelta(hours=24)).isoformat()
        recordings_dir = os.path.join(os.path.dirname(__file__), "..", "recordings")
        async with aio_connect() as db:
            db.row_factory = aiosqlite.Row
            async with db.execute(
                "SELECT id, filename, clip_count FROM recordings "
                "WHERE clipped = -1 AND skip_reason IS NULL AND transcribed = 2 "
                "AND start_time >= ? ORDER BY start_time DESC LIMIT 20",
                (cutoff,),
            ) as cur:
                rows = await cur.fetchall()
        if not rows:
            return
        queued = []
        for rec in rows:
            mp4_path = os.path.join(recordings_dir, rec["filename"])
            srt_path = resolve_srt_path(mp4_path)
            if not (os.path.exists(mp4_path) and srt_path):
                continue
            async with aio_connect() as db:
                await db.execute(
                    "UPDATE recordings SET clipped = 0, clip_error = NULL WHERE id = ?",
                    (rec["id"],),
                )
                await db.commit()
            asyncio.create_task(_run_editor(
                rec["id"], mp4_path, srt_path,
                clip_count=rec["clip_count"] or 1,
                broadcast_fn=broadcast,
            ))
            queued.append(rec["id"])
        if queued:
            logger.info(f"GPU online — auto-retrying {len(queued)} failed clip job(s): {queued}")
            await broadcast({"type": "gpu_auto_retry", "recording_ids": queued})

        # Also trigger director+creative for groups that have classic done but pipelines not run
        # Strategy: run director first, then creative only after director is done
        from transcribe import _run_director_pipeline, _run_creative_pipeline
        async with aio_connect() as db:
            db.row_factory = aiosqlite.Row
            # Director groups first
            async with db.execute(
                """SELECT id FROM clip_groups
                   WHERE classic_status = 2
                     AND (director_status IN (0, -1, -3))
                   ORDER BY id DESC"""
            ) as cur:
                pending_director = [r["id"] for r in await cur.fetchall()]
            # Creative groups where director is done
            async with db.execute(
                """SELECT id FROM clip_groups
                   WHERE classic_status = 2
                     AND director_status = 2
                     AND (creative_status IN (0, -1, -3) OR creative_status IS NULL)
                   ORDER BY id DESC"""
            ) as cur:
                pending_creative = [r["id"] for r in await cur.fetchall()]
        if pending_director:
            logger.info(f"GPU online — triggering director for {len(pending_director)} groups")
            for gid in pending_director:
                asyncio.create_task(_run_director_pipeline(gid))
                await asyncio.sleep(0.2)
        if pending_creative:
            logger.info(f"GPU online — triggering creative for {len(pending_creative)} groups (director done)")
            for gid in pending_creative:
                asyncio.create_task(_run_creative_pipeline(gid))
                await asyncio.sleep(0.2)
    except Exception as e:
        logger.warning(f"_on_gpu_online auto-retry failed: {e}")
    finally:
        # Always wake the transcription poll when GPU comes back online
        await flush_poll()


async def _periodic_cleanup():
    """Every 6 hours: clean up stale enhance jobs, GPU clip files, and old local recordings."""
    gpu_url = os.environ.get("GPU_SERVICE_URL", "http://10.190.0.203:8877")
    recordings_dir = os.path.join(os.path.dirname(__file__), "..", "recordings")
    while True:
        await asyncio.sleep(6 * 3600)
        try:
            # Remove all terminal-state enhance jobs (done/error/cancelled)
            stale = [
                jid for jid, j in list(_enhance_jobs.items())
                if j.get("status") in ("done", "error", "cancelled")
            ]
            for jid in stale:
                _enhance_jobs.pop(jid, None)
            if stale:
                logger.info(f"[cleanup] Removed {len(stale)} stale enhance jobs from memory")

            # Ask GPU server to delete completed clip output dirs
            async with httpx.AsyncClient(timeout=15.0) as client:
                resp = await client.post(f"{gpu_url}/maintenance/cleanup-clips")
                if resp.status_code == 200:
                    data = resp.json()
                    logger.info(f"[cleanup] GPU clips: deleted={data.get('deleted',0)} freed={data.get('freed_gb',0):.2f}GB")

            # Auto-delete local source MP4s for recordings whose publish task
            # completed more than 7 days ago (keeps disk from filling up).
            async with aio_connect() as db:
                db.row_factory = aiosqlite.Row
                async with db.execute(
                    """SELECT DISTINCT r.id, r.filename
                       FROM recordings r
                       JOIN publish_tasks pt ON pt.group_id = r.group_id
                       WHERE pt.status = 'done'
                         AND pt.published_at <= datetime('now', '-7 days')
                         AND r.local_deleted = 0
                         AND r.filename IS NOT NULL"""
                ) as cur:
                    old_recs = await cur.fetchall()

            deleted_ids = []
            deleted_count = 0
            for rec in old_recs:
                filepath = os.path.join(recordings_dir, rec["filename"])
                if os.path.exists(filepath):
                    try:
                        os.unlink(filepath)
                        deleted_count += 1
                    except OSError as e:
                        logger.warning(f"[cleanup] Could not delete {rec['filename']}: {e}")
                deleted_ids.append(rec["id"])

            if deleted_ids:
                placeholders = ",".join("?" * len(deleted_ids))
                async with aio_connect() as db:
                    await db.execute(
                        f"UPDATE recordings SET local_deleted = 1 WHERE id IN ({placeholders})",
                        deleted_ids,
                    )
                    await db.commit()
                logger.info(f"[cleanup] Auto-deleted {deleted_count} local source MP4s (published >7 days ago)")

            # Remove recording_clips rows whose files no longer exist on disk.
            # GPU cleanup can delete clip output dirs without cleaning up DB rows.
            async with aio_connect() as db:
                db.row_factory = aiosqlite.Row
                async with db.execute("SELECT id, clip_filename FROM recording_clips") as cur:
                    all_clips = await cur.fetchall()
            stale_rc_ids = [
                r["id"] for r in all_clips
                if r["clip_filename"] and not os.path.exists(
                    os.path.join(recordings_dir, r["clip_filename"])
                )
            ]
            if stale_rc_ids:
                placeholders = ",".join("?" * len(stale_rc_ids))
                async with aio_connect() as db:
                    await db.execute(
                        f"DELETE FROM recording_clips WHERE id IN ({placeholders})",
                        stale_rc_ids,
                    )
                    await db.commit()
                logger.info(f"[cleanup] Removed {len(stale_rc_ids)} stale recording_clips rows")

            await _cleanup_stale_open_recording_placeholders()

        except asyncio.CancelledError:
            return
        except Exception as e:
            logger.debug(f"Periodic cleanup error: {e}")


@asynccontextmanager
async def lifespan(app: FastAPI):
    # Deployment role is validated at module load (see _DEPLOYMENT_ROLE above)
    if not IS_CONTROL_PLANE:
        require_remote_gpu("backend startup")
    await init_db()
    logger.info("Qianchuan fact-source SQLite DB: %s", os.path.abspath(DB_PATH))
    await _reset_stuck_clip_tasks()
    await _cleanup_stale_open_recording_placeholders()
    # Load human-approved keyword score overrides into the scoring table
    from editor import load_rule_overrides
    await load_rule_overrides()

    from gpu_state import watch_gpu_service, register_online_callback
    gpu_watcher_task = asyncio.create_task(watch_gpu_service(broadcast_fn=broadcast))
    tasks = [gpu_watcher_task]

    if IS_CONTROL_PLANE:
        # The Mac instance is API/static-file control plane only.  In particular,
        # do not register the GPU-online callback: it dispatches media pipelines.
        logger.info(
            "DEPLOYMENT_ROLE=control-plane: background media jobs disabled "
            "(capture, transcription, backfill, publish, enhance, creative/director, and room monitors). "
            "Remote GPU backend: %s",
            REMOTE_BACKEND_URL,
        )
    else:
        register_online_callback(_on_gpu_online)
        logger.info(
            "DEPLOYMENT_ROLE=gpu-backend: media workers enabled "
            "(capture, transcription, backfill, publish, enhance, creative/director, and room monitors)"
        )
        tasks.extend([
            asyncio.create_task(poll_transcriptions(broadcast_fn=broadcast)),
            asyncio.create_task(backfill_auto_merge()),
            asyncio.create_task(poll_publish_tasks(broadcast_fn=broadcast)),
            asyncio.create_task(_memory_monitor(broadcast_fn=broadcast)),
            asyncio.create_task(_enhance_worker()),
            asyncio.create_task(_periodic_cleanup()),
            asyncio.create_task(_periodic_creative_dispatch()),
            asyncio.create_task(_periodic_director_dispatch()),
            asyncio.create_task(_periodic_qianchuan_dispatch()),
            asyncio.create_task(_periodic_clip_dispatch()),
        ])
        await monitor.start_all()
    yield
    for task in tasks:
        task.cancel()
    for task in tasks:
        try:
            await task
        except asyncio.CancelledError:
            pass
    for room_id in list(monitor._tasks.keys()):
        await monitor.remove_room(room_id)


APP_VERSION = "MVP1.04.2026032501"

# API loading state (module-level for health endpoint access)
_api_v2_loaded = False
_qianchuan_api_loaded = False
_qianchuan_api_error = None
_qianchuan_upload_loaded = False

app = FastAPI(title="Douyin Recorder", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/health")
async def health():
    """Health check with deployment role, remote backend status, and API routing info.

    On the control-plane (Mac, port 8099), this shows the remote GPU backend URL
    where media APIs (qianchuan, director, etc.) actually run.
    On the GPU backend (port 8899), this shows local service health.
    """
    result = {
        "ok": True,
        "deployment_role": _DEPLOYMENT_ROLE,
        "version": APP_VERSION,
        "remote_backend_url": REMOTE_BACKEND_URL,
    }
    if IS_CONTROL_PLANE:
        result["qianchuan_entry"] = f"{REMOTE_BACKEND_URL}/api/v2/qianchuan/status"
        result["director_entry"] = f"{REMOTE_BACKEND_URL}/api/v2/director/status"
        result["note"] = (
            "Control-plane: 本机仅提供轻量API代理和静态文件服务。"
            "千川投流版和导演模式API请使用远端GPU后端。"
        )
        # Probe remote backend liveness (best-effort, non-blocking)
        try:
            async with httpx.AsyncClient(timeout=5.0) as client:
                resp = await client.get(f"{REMOTE_BACKEND_URL}/")
                result["remote_backend_reachable"] = resp.status_code < 500
        except Exception:
            result["remote_backend_reachable"] = False
    else:
        # GPU backend: add qianchuan queue depth for operational visibility
        result["media_workers_enabled"] = True
        result["worker_services"] = [
            "transcription", "backfill", "publish", "enhance",
            "creative", "director", "qianchuan", "room-monitors",
        ]
        result["qianchuan_api_loaded"] = _qianchuan_api_loaded
        if "_qianchuan_api_error" in globals() and _qianchuan_api_error:
            result["qianchuan_api_error"] = _qianchuan_api_error
        try:
            async with aio_connect() as db:
                db.row_factory = aiosqlite.Row
                for status_code, label in [(0, "pending"), (1, "running"), (2, "done"), (-1, "failed"), (-2, "permanent_fail"), (-3, "quality_fail"), (-4, "probe_fail")]:
                    async with db.execute(
                        "SELECT COUNT(*) as cnt FROM clip_groups WHERE qianchuan_status = ?", (status_code,)
                    ) as cur:
                        row = await cur.fetchone()
                        result[f"qianchuan_{label}"] = row["cnt"]
        except Exception:
            pass
    return result


@app.get("/", include_in_schema=False)
async def service_entrypoint():
    """Open the production frontend at the canonical remote backend URL."""
    return RedirectResponse(url="/frontend/", status_code=307)


# ── API v2: Director & Qianchuan routes ─────────────────────────────────────
# Control-plane (Mac port 8099) gets lightweight proxy routes pointing to the
# remote GPU backend. GPU backend (port 8899) loads the full media pipeline.
if IS_CONTROL_PLANE:
    @app.get("/api/v2/qianchuan/status")
    async def qianchuan_status():
        return {
            "local_qianchuan_available": False,
            "deployment_role": "control-plane",
            "message": "千川投流版API运行在远端GPU后端（8899端口）",
            "remote_backend_url": REMOTE_BACKEND_URL,
            "remote_qianchuan_status_url": f"{REMOTE_BACKEND_URL}/api/v2/qianchuan/status",
        }

    @app.get("/api/v2/director/status")
    async def director_status():
        return {
            "local_director_available": False,
            "deployment_role": "control-plane",
            "message": "导演模式API运行在远端GPU后端（8899端口）",
            "remote_backend_url": REMOTE_BACKEND_URL,
            "remote_director_status_url": f"{REMOTE_BACKEND_URL}/api/v2/director/status",
        }

    logger.info("控制面: 已加载千川/导演模式代理路由 -> %s", REMOTE_BACKEND_URL)
else:
    # GPU backend: full api_v2 routes for media pipelines.
    # Each import is isolated so a single missing dependency doesn't cascade.
    _api_v2_loaded = False
    _qianchuan_api_loaded = False
    _qianchuan_api_error = None
    _qianchuan_upload_loaded = False

    # ── api_v2 (director + qianchuan generate/pipeline) ──
    try:
        from api_v2 import director_router, set_broadcast_fn
        app.include_router(director_router)
        set_broadcast_fn(broadcast)
        _api_v2_loaded = True
        logger.info("导演模式API路由已加载")
    except Exception as exc:
        logger.warning(f"导演模式API加载失败: {exc}")

    # Keep Qianchuan registration independent from the director stack.  A
    # director-only dependency failure must not silently remove the remote
    # GPU backend's Qianchuan API.
    try:
        from api_v2 import qianchuan_router, set_broadcast_fn
        app.include_router(qianchuan_router)
        set_broadcast_fn(broadcast)
        _qianchuan_api_loaded = True
        logger.info("千川投流版API路由已加载")
    except Exception as exc:
        _qianchuan_api_error = f"{type(exc).__name__}: {exc}"
        logger.warning(f"千川投流版API加载失败: {_qianchuan_api_error}")

    # ── qianchuan_upload (learning material upload) ──
    try:
        from qianchuan_upload import qianchuan_upload_router
        app.include_router(qianchuan_upload_router)
        _qianchuan_upload_loaded = True
        logger.info("千川投流版素材上传API路由已加载")
    except Exception as exc:
        logger.warning(f"千川投流版上传API加载失败: {exc}")

    # ── Fallback status endpoints when modules are unavailable ──
    if not _api_v2_loaded:
        @app.get("/api/v2/director/status")
        async def _fallback_director_status():
            return {
                "director_available": False,
                "deployment_role": "gpu-backend",
                "error": "导演模式API模块加载失败 — 检查 backend/api_v2.py 及其依赖项",
            }

    if not _qianchuan_api_loaded:
        @app.get("/api/v2/qianchuan/status")
        async def _fallback_qianchuan_status():
            return {
                "qianchuan_available": False,
                "deployment_role": "gpu-backend",
                "error": _qianchuan_api_error or "千川投流版API模块加载失败 — 检查 backend/api_v2.py 及其依赖项",
                "queue_depth": 0,
                "running": 0,
                "done": 0,
                "failed": 0,
            }


# ── Rooms ────────────────────────────────────────────────────────────────────

@app.get("/api/monitor/status")
async def monitor_status():
    """Return room monitor lifecycle and operational health."""
    return await monitor.get_overall_status()


@app.post("/api/monitor/start")
async def monitor_start():
    """Start enabled room monitors. Idempotent for already-running rooms."""
    if IS_CONTROL_PLANE:
        return {"ok": False, "running": False, "reason": "media monitors run on the GPU backend"}
    await monitor.start_all()
    return {"ok": True, "status": await monitor.get_overall_status()}


@app.post("/api/monitor/restart")
async def monitor_restart():
    """Stop and start the room monitor service."""
    if IS_CONTROL_PLANE:
        return {"ok": False, "running": False, "reason": "media monitors run on the GPU backend"}
    await monitor.restart()
    return {"ok": True, "status": await monitor.get_overall_status()}

@app.get("/api/rooms")
async def list_rooms():
    global _rooms_cache
    import time
    now = time.monotonic()
    if _rooms_cache and now - _rooms_cache[0] < _API_CACHE_TTL_SECONDS:
        rows_data = _rooms_cache[1]
    else:
        async with aio_connect() as db:
            db.row_factory = aiosqlite.Row
            async with db.execute("SELECT * FROM rooms ORDER BY created_at DESC") as cur:
                rows_data = [dict(row) for row in await cur.fetchall()]
        _rooms_cache = (now, rows_data)
    result = []
    for r in rows_data:
        status = monitor.get_status(r["id"])
        result.append({
            "id": r["id"],
            "name": r["name"],
            "url": r["url"],
            "enabled": bool(r["enabled"]),
            "created_at": r["created_at"],
            **status,
        })
    return result


@app.post("/api/rooms", status_code=201)
async def add_room(body: RoomCreate):
    global _rooms_cache
    try:
        async with aio_connect() as db:
            cur = await db.execute(
                "INSERT INTO rooms (name, url) VALUES (?, ?)",
                (body.name, body.url)
            )
            await db.commit()
            room_id = cur.lastrowid
        await monitor.add_room(room_id, body.name, body.url)
        _rooms_cache = None
        return {"id": room_id, "name": body.name, "url": body.url, "enabled": True}
    except aiosqlite.IntegrityError:
        raise HTTPException(status_code=409, detail="Room URL already exists")


@app.delete("/api/rooms/{room_id}", status_code=204)
async def delete_room(room_id: int):
    global _rooms_cache
    async with aio_connect() as db:
        await db.execute("DELETE FROM rooms WHERE id = ?", (room_id,))
        await db.commit()
    await monitor.remove_room(room_id)
    _rooms_cache = None


@app.patch("/api/rooms/{room_id}/toggle")
async def toggle_room(room_id: int):
    global _rooms_cache
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM rooms WHERE id = ?", (room_id,)) as cur:
            room = await cur.fetchone()
        if not room:
            raise HTTPException(status_code=404, detail="Room not found")
        new_enabled = 0 if room["enabled"] else 1
        await db.execute("UPDATE rooms SET enabled = ? WHERE id = ?", (new_enabled, room_id))
        await db.commit()

    if new_enabled:
        async with aio_connect() as db:
            db.row_factory = aiosqlite.Row
            async with db.execute("SELECT * FROM rooms WHERE id = ?", (room_id,)) as cur:
                room = await cur.fetchone()
        await monitor.add_room(room["id"], room["name"], room["url"])
    else:
        await monitor.remove_room(room_id)

    _rooms_cache = None

    return {"id": room_id, "enabled": bool(new_enabled)}


# ── Recordings ───────────────────────────────────────────────────────────────

async def _upload_gpu_then_edit(recording_id: int, filepath: str, srt_path: str, room_id: int, clip_duration, clip_count: int):
    """Upload MP4 to GPU server so clip-jobs can find it, then run the editor."""
    try:
        await sync_file(filepath, room_id)
    except Exception as e:
        logger.warning(f"GPU pre-upload failed for recording {recording_id}: {e}")
    await _run_editor(recording_id, filepath, srt_path, clip_duration=clip_duration, clip_count=clip_count, broadcast_fn=broadcast)


@app.post("/api/rooms/{room_id}/upload", status_code=201)
async def upload_recording(room_id: int, file: UploadFile = File(...), srt: Optional[UploadFile] = File(None), duration_sec: Optional[float] = Form(None), clip_count: int = Form(1)):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT id FROM rooms WHERE id = ?", (room_id,)) as cur:
            if not await cur.fetchone():
                raise HTTPException(status_code=404, detail="Room not found")

    clip_count = max(1, min(5, clip_count))

    now = datetime.utcnow()
    ts = now.strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r"[^\w\-.]", "_", file.filename or "upload.mp4")
    filename = f"{ts}_{safe_name}"
    recordings_dir = os.path.join(os.path.dirname(__file__), "..", "recordings")
    os.makedirs(recordings_dir, exist_ok=True)
    filepath = os.path.join(recordings_dir, filename)

    # Stream to disk in 1 MB chunks to avoid loading large video files into RAM
    size_bytes = 0
    with open(filepath, "wb") as f:
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            f.write(chunk)
            size_bytes += len(chunk)

    start_time = now.isoformat()
    async with aio_connect() as db:
        cur = await db.execute(
            "INSERT INTO recordings (room_id, filename, start_time, end_time, size_bytes, synced, clip_count) VALUES (?, ?, ?, ?, ?, 0, ?)",
            (room_id, filename, start_time, start_time, size_bytes, clip_count),
        )
        await db.commit()
        recording_id = cur.lastrowid

    asyncio.create_task(_generate_upload_thumb(recording_id, filepath))

    # If SRT is provided, skip GPU transcription and trigger clipping immediately
    if srt is not None:
        srt_filename = os.path.splitext(filename)[0] + ".srt"
        srt_path = os.path.join(recordings_dir, srt_filename)
        srt_content = await srt.read()
        with open(srt_path, "wb") as f:
            f.write(srt_content)
        async with aio_connect() as db:
            await db.execute(
                "UPDATE recordings SET transcribed = 2, synced = 1 WHERE id = ?",
                (recording_id,),
            )
            await db.commit()
        # Upload MP4 to GPU server first so clip-jobs can find the file, then edit
        asyncio.create_task(_upload_gpu_then_edit(recording_id, filepath, srt_path, room_id, duration_sec, clip_count))
        return {"id": recording_id, "filename": filename, "size_bytes": size_bytes, "gpu_job_id": None}

    from comfyui_client import free_vram
    await free_vram()
    job_id = await sync_file(filepath, room_id)
    if job_id:
        async with aio_connect() as db:
            await db.execute(
                "UPDATE recordings SET transcribed = 1, synced = 1, gpu_job_id = ? WHERE id = ?",
                (job_id, recording_id),
            )
            await db.commit()
    else:
        logger.warning(f"Upload accepted but GPU sync failed for {filename}")

    return {"id": recording_id, "filename": filename, "size_bytes": size_bytes, "gpu_job_id": job_id}


async def _generate_upload_thumb(recording_id: int, mp4_path: str):
    thumb = await generate_thumbnail(mp4_path, offset=5.0)
    if thumb:
        async with aio_connect() as db:
            await db.execute(
                "UPDATE recordings SET thumbnail = ? WHERE id = ?",
                (os.path.relpath(thumb, os.path.join(os.path.dirname(__file__), "..", "recordings")), recording_id),
            )
            await db.commit()


@app.get("/api/rooms/{room_id}/recordings")
async def list_recordings(room_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT * FROM recordings WHERE room_id = ? ORDER BY start_time DESC LIMIT 500",
            (room_id,)
        ) as cur:
            rows = await cur.fetchall()
    return [dict(r) for r in rows]


_STATUS_WHERE = {
    "transcribe_running": "r.transcribed = 1",
    "transcribe_pending": "r.transcribed = 0 AND r.local_deleted = 0 AND r.end_time IS NOT NULL",
    "transcribe_failed":  "r.transcribed = -1 AND (r.skip_reason IS NULL OR r.skip_reason = '')",
    "clip_running":       "r.clipped = 1",
    "clip_pending":       "r.transcribed = 2 AND r.clipped = 0",
    "clip_failed":        "r.clipped = -1 AND (r.skip_reason IS NULL OR r.skip_reason != '已手动清除')",
    "running":            "(r.transcribed = 1 OR r.clipped = 1)",
    # top-level filter bar shortcuts
    "success":  "r.clipped = 2",
    "failed":   "((r.transcribed = -1 OR r.clipped = -1) AND (r.skip_reason IS NULL OR r.skip_reason = ''))",
    "active":   "(r.transcribed = 1 OR r.clipped = 1)",
}

_SORT_COLS = {
    "start_time": "r.start_time",
    "filename":   "r.filename",
    "id":         "r.id",
}


@app.get("/api/recordings")
async def list_all_recordings(
    page: int = 1,
    limit: int = Query(default=50, ge=1, le=500),
    status: Optional[str] = None,
    sort: str = "start_time",
    order: str = "desc",
):
    page = max(1, page)
    offset = (page - 1) * limit
    where = f"WHERE {_STATUS_WHERE[status]}" if status in _STATUS_WHERE else ""
    col = _SORT_COLS.get(sort, "r.start_time")
    direction = "ASC" if order.lower() == "asc" else "DESC"
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(f"SELECT COUNT(*) FROM recordings r {where}") as cur:
            (total,) = await cur.fetchone()
        async with db.execute(f"""
            SELECT r.*, rm.name as room_name
            FROM recordings r
            JOIN rooms rm ON r.room_id = rm.id
            {where}
            ORDER BY {col} {direction}
            LIMIT ? OFFSET ?
        """, (limit, offset)) as cur:
            rows = await cur.fetchall()
    return {
        "items": [dict(r) for r in rows],
        "total": total,
        "page": page,
        "pages": max(1, (total + limit - 1) // limit),
    }


@app.get("/api/recording-clips/bulk")
async def bulk_recording_clips(ids: str = ""):
    """Return all clips for a comma-separated list of recording ids."""
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    if not id_list:
        return {}
    placeholders = ",".join("?" * len(id_list))
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        # Only return the latest retry's clips (highest id per recording+variant)
        async with db.execute(
            f"""SELECT rc.* FROM recording_clips rc
                INNER JOIN (
                    SELECT recording_id, variant_idx, MAX(id) as max_id
                    FROM recording_clips
                    WHERE recording_id IN ({placeholders})
                    GROUP BY recording_id, variant_idx
                ) latest ON rc.id = latest.max_id
                ORDER BY rc.recording_id, rc.variant_idx ASC""",
            id_list,
        ) as cur:
            rows = await cur.fetchall()
    result: dict = {}
    for r in rows:
        result.setdefault(str(r["recording_id"]), []).append(dict(r))
    return result


@app.get("/api/clips")
async def list_clips():
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("""
            SELECT r.id, r.filename, r.clip_filename, r.start_time, r.end_time,
                   r.room_id, rm.name as room_name, rc.gpu_clip_job_id
            FROM recordings r
            JOIN rooms rm ON r.room_id = rm.id
            LEFT JOIN recording_clips rc ON rc.recording_id = r.id AND rc.variant_idx = 0
            WHERE r.clipped = 2 AND r.clip_filename IS NOT NULL
            ORDER BY r.start_time DESC
        """) as cur:
            rows = await cur.fetchall()
    
    # Fetch clip sizes from GPU service and scan local clips directory
    clip_size_map = {}
    
    # Try GPU service API first
    try:
        gpu_url = os.environ.get("GPU_SERVICE_URL", "http://10.190.0.203:8877")
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get(f"{gpu_url}/clip-jobs")
            if resp.status_code == 200:
                gpu_jobs = resp.json()
                for gj in gpu_jobs:
                    # Map by job_id (which is the hash)
                    if gj.get("job_id"):
                        clip_size_map[gj["job_id"]] = gj.get("size")
                    # Map by output_path hash
                    if gj.get("output_path"):
                        parts = gj["output_path"].replace("\\", "/").split("/")
                        if len(parts) >= 3:
                            clip_size_map[parts[-2]] = gj.get("size")  # hash dir name
    except Exception as e:
        logger.warning(f"Failed to fetch clip sizes from GPU service: {e}")
    
    # Also scan local clips directory as fallback
    # Try multiple possible locations for Windows
    _script_dir = os.path.dirname(__file__)
    _base_dirs = [
        os.path.join(_script_dir, "..", "douyin_recordings", "clips"),
        r"C:\Users\neo\douyin_recordings\clips",
        r"C:\douyin_recordings\clips",
    ]
    clips_dir = os.environ.get("CLIPS_DIR", _base_dirs[0])
    if not os.path.exists(clips_dir):
        for d in _base_dirs:
            if os.path.exists(d):
                clips_dir = d
                break
    try:
        if os.path.exists(clips_dir):
            for clip_file in Path(clips_dir).rglob("clip.mp4"):
                if clip_file.is_file():
                    hash_dir = clip_file.parent.name
                    clip_size_map[hash_dir] = clip_file.stat().st_size
    except Exception as e:
        logger.warning(f"Failed to scan clips directory: {e}")
    
    # Build filename→size mapping from GPU jobs (match by basename)
    basename_size_map = {}
    for gj in gpu_jobs:
        out = gj.get("output_path", "")
        size = gj.get("size")
        if size and out:
            bn = os.path.basename(out)  # e.g. "clip.mp4"
            if bn not in basename_size_map:
                basename_size_map[bn] = size

    items = []
    for r in rows:
        clip_filename = r["clip_filename"]
        gpu_job_id = r["gpu_clip_job_id"]
        size = None

        # 1. Direct hash-dir match from GPU jobs
        size = clip_size_map.get(clip_filename) or clip_size_map.get(gpu_job_id)

        # 2. MD5 hash of clip_filename → clips/<hash>/clip.mp4
        if size is None and clip_filename:
            _hash = hashlib.md5(clip_filename.encode()).hexdigest()[:12]
            for sep in ["/", "\\"]:
                hash_path = f"clips{sep}{_hash}{sep}clip.mp4"
                if hash_path in clip_size_map:
                    size = clip_size_map[hash_path]
                    break

        # 3. Match by basename (GPU stores clips/<md5>/clip.mp4, DB stores room/date/filename_clip.mp4)
        if size is None and clip_filename:
            bn = os.path.basename(clip_filename)
            if bn in basename_size_map:
                size = basename_size_map[bn]

        # 4. Fallback: check local path
        if size is None and clip_filename:
            clip_path = os.path.abspath(
                os.path.join(os.path.dirname(__file__), "..", "recordings", clip_filename)
            )
            size = os.path.getsize(clip_path) if os.path.exists(clip_path) else None

        items.append({**dict(r), "clip_size": size})
    return items


# ── Subtitles ────────────────────────────────────────────────────────────────

@app.get("/api/recordings/{recording_id}/clip")
async def download_clip(recording_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM recordings WHERE id = ?", (recording_id,)) as cur:
            rec = await cur.fetchone()
    if not rec:
        raise HTTPException(status_code=404, detail="Recording not found")
    if rec["clipped"] != 2 or not rec["clip_filename"]:
        raise HTTPException(status_code=404, detail="Clip not ready")
    clip_path = os.path.join(os.path.dirname(__file__), "..", "recordings", rec["clip_filename"])
    if not os.path.exists(clip_path):
        raise HTTPException(status_code=404, detail="Clip file missing")
    return FileResponse(clip_path, media_type="video/mp4", filename=os.path.basename(rec["clip_filename"]))


@app.get("/api/recordings/{recording_id}/clips")
async def list_recording_clips(recording_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT * FROM recording_clips WHERE recording_id = ? ORDER BY variant_idx ASC",
            (recording_id,)
        ) as cur:
            rows = await cur.fetchall()
    return [dict(r) for r in rows]


@app.get("/api/recording-clips/{clip_id}/download")
async def download_recording_clip(clip_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM recording_clips WHERE id = ?", (clip_id,)) as cur:
            clip = await cur.fetchone()
    if not clip:
        raise HTTPException(status_code=404, detail="Clip not found")
    clip_path = os.path.join(os.path.dirname(__file__), "..", "recordings", clip["clip_filename"])
    if not os.path.exists(clip_path):
        raise HTTPException(status_code=404, detail="Clip file missing")
    return FileResponse(clip_path, media_type="video/mp4", filename=os.path.basename(clip["clip_filename"]))


@app.post("/api/recordings/{recording_id}/reclip", status_code=200)
async def reclip_recording(recording_id: int, body: dict = Body({})):
    """Reset a completed clip and re-run the editor, optionally with user feedback."""
    feedback = (body.get("feedback") or "").strip() or None
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT * FROM recordings WHERE id=?", (recording_id,)
        ) as cur:
            rec = await cur.fetchone()
    if not rec:
        raise HTTPException(status_code=404, detail="Recording not found")
    if rec["transcribed"] != 2:
        raise HTTPException(status_code=409, detail="Transcription not complete yet")

    # Delete old clip files for this recording
    old_clip = rec["clip_filename"]
    if old_clip:
        old_path = os.path.join(os.path.dirname(__file__), "..", "recordings", old_clip)
        try:
            if os.path.exists(old_path):
                os.unlink(old_path)
        except Exception as e:
            logger.warning(f"Could not delete old clip {old_clip}: {e}")

    # Delete old recording_clips rows
    async with aio_connect() as db:
        await db.execute("DELETE FROM recording_clips WHERE recording_id=?", (recording_id,))
        await db.execute(
            "UPDATE recordings SET clipped=0, clip_filename=NULL, thumbnail=NULL, reclip_feedback=? WHERE id=?",
            (feedback, recording_id),
        )
        await db.commit()

    # Re-trigger the editor with feedback
    mp4_path = os.path.join(os.path.dirname(__file__), "..", "recordings", rec["filename"])
    srt_path = resolve_srt_path(mp4_path)
    if not srt_path:
        raise HTTPException(status_code=404, detail="Non-empty SRT file missing")
    asyncio.create_task(
        _run_editor(recording_id, mp4_path, srt_path,
                    clip_count=rec["clip_count"] or 1,
                    feedback=feedback,
                    broadcast_fn=broadcast)
    )
    return {"ok": True, "recording_id": recording_id, "feedback": feedback}


@app.get("/api/recordings/{recording_id}/thumbnail")
async def get_thumbnail(recording_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT thumbnail FROM recordings WHERE id = ?", (recording_id,)) as cur:
            rec = await cur.fetchone()
    if not rec:
        raise HTTPException(status_code=404, detail="Recording not found")
    if not rec["thumbnail"]:
        raise HTTPException(status_code=404, detail="Thumbnail not available")
    thumb_path = os.path.join(os.path.dirname(__file__), "..", "recordings", rec["thumbnail"])
    if not os.path.exists(thumb_path):
        raise HTTPException(status_code=404, detail="Thumbnail file missing")
    return FileResponse(thumb_path, media_type="image/jpeg")


@app.get("/api/recordings/{recording_id}/srt")
async def download_srt(recording_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM recordings WHERE id = ?", (recording_id,)) as cur:
            rec = await cur.fetchone()
    if not rec:
        raise HTTPException(status_code=404, detail="Recording not found")
    if rec["transcribed"] != 2:
        raise HTTPException(status_code=404, detail="Subtitle not ready")
    srt_path = resolve_srt_path(os.path.join(os.path.dirname(__file__), "..", "recordings", rec["filename"]))
    if not srt_path:
        raise HTTPException(status_code=404, detail="SRT file missing")
    return FileResponse(srt_path, media_type="text/plain", filename=os.path.basename(srt_path))


# ── Human Review & Learning System ───────────────────────────────────────────

@app.get("/api/recordings/{recording_id}/review-candidates")
async def get_review_candidates(recording_id: int):
    """Return all scored segments from the SRT for human review.

    Segments are returned as a flat list with idx/text/start/end/score/category/valid.
    If review_candidates is already cached in DB, return it directly.
    """
    from editor import parse_srt, score_and_tag, _merge_short_segs

    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM recordings WHERE id = ?", (recording_id,)) as cur:
            rec = await cur.fetchone()
    if not rec:
        raise HTTPException(status_code=404, detail="Recording not found")
    if rec["transcribed"] != 2:
        raise HTTPException(status_code=409, detail="Transcription not ready")

    # Build segments (recompute fresh each call to reflect latest _SCORES_EFFECTIVE)
    srt_path = resolve_srt_path(os.path.join(os.path.dirname(__file__), "..", "recordings", rec["filename"]))
    if not srt_path:
        raise HTTPException(status_code=404, detail="SRT file missing")

    raw_segs = parse_srt(srt_path)
    segs = _merge_short_segs(raw_segs)
    for seg in segs:
        score_and_tag(seg)

    all_segs = [
        {
            "idx":      seg.idx,
            "start":    round(seg.start, 2),
            "end":      round(seg.end, 2),
            "text":     seg.text,
            "score":    round(getattr(seg, "score", 0.0), 2),
            "category": getattr(seg, "category", ""),
            "valid":    getattr(seg, "valid", True),
        }
        for seg in segs
    ]
    result = {"all_segs": all_segs, "total": len(all_segs)}

    # Cache in DB
    async with aio_connect() as db:
        await db.execute(
            "UPDATE recordings SET review_candidates=? WHERE id=?",
            (json.dumps(result, ensure_ascii=False), recording_id)
        )
        await db.commit()

    # Attach latest clip_review data if exists
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT algo_segments, user_segments, user_added, user_removed FROM clip_reviews WHERE recording_id=? ORDER BY id DESC LIMIT 1",
            (recording_id,)
        ) as cur:
            existing = await cur.fetchone()

    if existing:
        result["prev_review"] = {
            "algo_segments":  json.loads(existing["algo_segments"]  or "[]"),
            "user_segments":  json.loads(existing["user_segments"]  or "[]"),
            "user_added":     json.loads(existing["user_added"]     or "[]"),
            "user_removed":   json.loads(existing["user_removed"]   or "[]"),
        }
    result["review_status"] = rec["review_status"] or 0

    return result


@app.post("/api/recordings/{recording_id}/review", status_code=200)
async def submit_review(recording_id: int, body: dict = Body({})):
    """Save human review result and trigger training cycle.

    body: {
        algo_segments: [idx, ...],   # indices the algo selected
        user_segments: [idx, ...],   # indices user kept after review
        user_added:    [idx, ...],   # indices user added (algo missed)
        user_removed:  [idx, ...],   # indices user removed (algo wrong)
        user_segments_full: [{idx, text, start, end, ...}, ...]  # full segment data for added
    }
    """
    from rule_trainer import run_training_cycle

    algo_segs    = body.get("algo_segments", [])
    user_segs    = body.get("user_segments", [])
    user_added   = body.get("user_added", [])
    user_removed = body.get("user_removed", [])
    user_segs_full = body.get("user_segments_full", [])

    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT id FROM recordings WHERE id=?", (recording_id,)) as cur:
            if not await cur.fetchone():
                raise HTTPException(status_code=404, detail="Recording not found")

        # Upsert: replace any existing review for this recording
        await db.execute("DELETE FROM clip_reviews WHERE recording_id=?", (recording_id,))
        await db.execute("""
            INSERT INTO clip_reviews
                (recording_id, algo_segments, user_segments, user_added, user_removed, is_valid_sample, user_segments_full)
            VALUES (?, ?, ?, ?, ?, 1, ?)
        """, (
            recording_id,
            json.dumps(algo_segs),
            json.dumps(user_segs),
            json.dumps(user_added),
            json.dumps(user_removed),
            json.dumps(user_segs_full, ensure_ascii=False),
        ))
        # Mark review_status=1 (reviewed)
        await db.execute("UPDATE recordings SET review_status=1 WHERE id=?", (recording_id,))
        await db.commit()

    # Trigger training asynchronously (don't block the response)
    asyncio.create_task(run_training_cycle())

    return {"ok": True, "recording_id": recording_id, "user_kept": len(user_segs)}


@app.get("/api/rule-suggestions")
async def list_rule_suggestions(status: str = "pending"):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT * FROM rule_suggestions WHERE status=? ORDER BY created_at DESC",
            (status,)
        ) as cur:
            rows = await cur.fetchall()
    return [dict(r) for r in rows]


@app.post("/api/rule-suggestions/{suggestion_id}/accept", status_code=200)
async def accept_rule_suggestion(suggestion_id: int):
    """Accept a suggestion: write to rule_overrides and hot-reload _SCORES_EFFECTIVE."""
    from editor import _SCORES_EFFECTIVE

    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM rule_suggestions WHERE id=?", (suggestion_id,)) as cur:
            sug = await cur.fetchone()
        if not sug:
            raise HTTPException(status_code=404, detail="Suggestion not found")
        if sug["status"] != "pending":
            raise HTTPException(status_code=409, detail=f"Already {sug['status']}")

        # Write / update rule_overrides
        await db.execute("""
            INSERT INTO rule_overrides (keyword, score, note, updated_at)
            VALUES (?, ?, ?, datetime('now'))
            ON CONFLICT(keyword) DO UPDATE SET
                score=excluded.score,
                note=excluded.note,
                updated_at=datetime('now')
        """, (sug["keyword"], sug["suggested_score"], sug["reason"]))

        await db.execute(
            "UPDATE rule_suggestions SET status='accepted', resolved_at=datetime('now') WHERE id=?",
            (suggestion_id,)
        )
        await db.commit()

    # Hot-reload _SCORES_EFFECTIVE
    _SCORES_EFFECTIVE[sug["keyword"]] = sug["suggested_score"]
    logger.info(f"rule_override accepted: {sug['keyword']} → {sug['suggested_score']}")

    return {"ok": True, "keyword": sug["keyword"], "new_score": sug["suggested_score"]}


@app.post("/api/rule-suggestions/{suggestion_id}/reject", status_code=200)
async def reject_rule_suggestion(suggestion_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT id, status FROM rule_suggestions WHERE id=?", (suggestion_id,)) as cur:
            sug = await cur.fetchone()
        if not sug:
            raise HTTPException(status_code=404, detail="Suggestion not found")
        if sug["status"] != "pending":
            raise HTTPException(status_code=409, detail=f"Already {sug['status']}")
        await db.execute(
            "UPDATE rule_suggestions SET status='rejected', resolved_at=datetime('now') WHERE id=?",
            (suggestion_id,)
        )
        await db.commit()
    return {"ok": True}


# ── Groups ───────────────────────────────────────────────────────────────────

@app.get("/api/groups")
async def list_groups():
    global _groups_cache
    import time
    now = time.monotonic()
    if _groups_cache and now - _groups_cache[0] < _API_CACHE_TTL_SECONDS:
        return _groups_cache[1]
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("""
            SELECT g.*,
                   rm.name as room_name,
                   COUNT(DISTINCT r.id) as clip_count,
                   SUM(CASE WHEN r.clipped = 2 THEN 1 ELSE 0 END) as ready_count,
                   COUNT(DISTINCT CASE WHEN pt.status IN ('done','publishing','pending','scheduled') THEN pt.id END) as published_count
            FROM clip_groups g
            LEFT JOIN rooms rm ON g.room_id = rm.id
            LEFT JOIN recordings r ON r.group_id = g.id
            LEFT JOIN publish_tasks pt ON pt.group_id = g.id
            GROUP BY g.id
            ORDER BY g.created_at DESC
        """) as cur:
            rows = await cur.fetchall()
    result = []
    for row in rows:
        item = dict(row) | _artifact_statuses(dict(row))
        item.update(_publish_version_metadata(item))
        result.append(item)
    _groups_cache = (now, result)
    return result


def _artifact_statuses(group: dict) -> dict:
    statuses = {}
    for version, field in (("classic", "merged_filename"), ("director", "director_final_video"),
                           ("realistic", "realistic_final_video"), ("conservative", "conservative_final_video"),
                           ("creative", "creative_final_video"), ("qianchuan", "qianchuan_final_video")):
        path, reason = resolve_artifact_path(group.get(field), version)
        statuses[f"{version}_file_status"] = "ready" if path else reason
        statuses[f"{version}_available"] = bool(path)
    return statuses


PUBLISHABLE_VERSIONS = ("classic", "director", "realistic", "conservative", "qianchuan")
PUBLISH_VERSION_LABELS = {
    "classic": "经典版",
    "director": "导演版",
    "realistic": "直出版",
    "conservative": "保守版",
    "qianchuan": "千川版",
}


def _publish_version_metadata(group: dict) -> dict:
    """Expose one stable, frontend-friendly availability object per publish version."""
    download_paths = {
        "classic": "download",
        "director": "director-download",
        "realistic": "realistic-download",
        "conservative": "conservative-download",
        "qianchuan": "qianchuan-download",
    }
    return {
        "publish_versions": [
            {
                "version": version,
                "label": PUBLISH_VERSION_LABELS[version],
                "available": bool(group.get(f"{version}_available")),
                "status": group.get(f"{version}_file_status", "not_generated"),
                "download_url": f"/api/groups/{group['id']}/{download_paths[version]}",
            }
            for version in PUBLISHABLE_VERSIONS
        ]
    }

async def get_group(group_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT g.*, rm.name as room_name FROM clip_groups g LEFT JOIN rooms rm ON g.room_id = rm.id WHERE g.id = ?",
            (group_id,)
        ) as cur:
            group = await cur.fetchone()
        if not group:
            raise HTTPException(status_code=404, detail="Group not found")
        async with db.execute(
            """SELECT id, filename, clip_filename, thumbnail, start_time, end_time,
                      session_label, has_tryon, has_promotion, transcribed, clipped, transcribe_error
               FROM recordings WHERE group_id = ? ORDER BY start_time ASC""",
            (group_id,)
        ) as cur:
            recs = await cur.fetchall()
    group_data = dict(group) | _artifact_statuses(dict(group))
    group_data.update(_publish_version_metadata(group_data))
    return group_data | {"recordings": [dict(r) for r in recs]}


@app.post("/api/groups/{group_id}/merge")
async def trigger_merge(group_id: int, force: bool = False):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM clip_groups WHERE id = ?", (group_id,)) as cur:
            group = await cur.fetchone()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    if not force and group["classic_status"] == 1 and group["director_status"] == 1 and (group["creative_status"] or 0) == 1:
        raise HTTPException(status_code=409, detail="All pipelines already in progress")
    # Reset all three pipelines and clear errors; force=True resets even completed ones
    async with aio_connect() as db:
        if force:
            await db.execute(
                """UPDATE clip_groups SET
                   quality_issue = NULL, director_error = NULL, merge_error = NULL, creative_error = NULL,
                   classic_status = 0, director_status = 0, creative_status = 0,
                   merged_filename = NULL, director_final_video = NULL, creative_final_video = NULL,
                   merge_status = 0, merged_at = NULL
                   WHERE id = ?""",
                (group_id,)
            )
        else:
            await db.execute(
                """UPDATE clip_groups SET
                   quality_issue = NULL, director_error = NULL, merge_error = NULL, creative_error = NULL,
                   classic_status  = 0,
                   director_status = 0,
                   creative_status = 0
                   WHERE id = ?""",
                (group_id,)
            )
        await db.commit()
    from transcribe import _run_director_pipeline, _run_creative_pipeline, _run_variant_pipeline
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT COUNT(*) as done FROM recordings WHERE group_id = ? AND clipped = 2",
            (group_id,),
        ) as cur:
            row = await cur.fetchone()
    valid_clips = row["done"] if row else 0
    if valid_clips == 0:
        async with aio_connect() as db:
            await db.execute(
                "UPDATE clip_groups SET classic_status = -1, merge_error = ? WHERE id = ?",
                ("无有效剪辑片段，无法合并", group_id),
            )
            await db.commit()
        logger.warning(f"Group {group_id}: no valid clips, skipping classic merge")
    else:
        asyncio.create_task(merge_group(group_id))
    asyncio.create_task(_run_director_pipeline(group_id))
    asyncio.create_task(_run_creative_pipeline(group_id))
    asyncio.create_task(_run_variant_pipeline(group_id, "realistic"))
    asyncio.create_task(_run_variant_pipeline(group_id, "conservative"))
    return {"group_id": group_id, "merge_status": 1}


@app.post("/api/groups/{group_id}/retry-modes")
async def retry_director_creative(group_id: int):
    """Re-queue director + creative pipelines only (classic merge not touched)."""
    async with aio_connect() as db:
        await db.execute(
            """UPDATE clip_groups SET
               director_status = 0, director_error = NULL,
               creative_status = 0, creative_error = NULL,
               director_final_video = NULL, creative_final_video = NULL
               WHERE id = ?""",
            (group_id,),
        )
        await db.commit()
    from transcribe import _run_director_pipeline, _run_creative_pipeline, _run_variant_pipeline
    asyncio.create_task(_run_director_pipeline(group_id))
    asyncio.create_task(_run_creative_pipeline(group_id))
    asyncio.create_task(_run_variant_pipeline(group_id, "realistic"))
    asyncio.create_task(_run_variant_pipeline(group_id, "conservative"))
    return {"group_id": group_id, "status": "queued"}


@app.patch("/api/groups/{group_id}/publish-versions")
async def set_publish_versions(group_id: int, body: dict):
    versions = body.get("publish_versions", "both")
    if versions not in ("classic", "director", "creative", "realistic", "conservative", "qianchuan", "both"):
        raise HTTPException(status_code=400, detail="publish_versions must be one of classic, director, creative, realistic, conservative, qianchuan, or both")
    async with aio_connect() as db:
        await db.execute(
            "UPDATE clip_groups SET publish_versions = ? WHERE id = ?", (versions, group_id)
        )
        await db.commit()
    return {"group_id": group_id, "publish_versions": versions}


@app.get("/api/groups/{group_id}/download")
async def download_merged(group_id: int, request: Request):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM clip_groups WHERE id = ?", (group_id,)) as cur:
            group = await cur.fetchone()
        if not group:
            raise HTTPException(status_code=404, detail="Group not found")
        # Prefer merged video; fall back to any ready clip in the group
        if group["merge_status"] == 2 and group["merged_filename"]:
            rel_path = group["merged_filename"]
        else:
            async with db.execute(
                "SELECT clip_filename FROM recordings WHERE group_id = ? AND clip_filename IS NOT NULL AND clipped = 2 ORDER BY id DESC LIMIT 1",
                (group_id,),
            ) as cur:
                rec = await cur.fetchone()
            if not rec:
                raise HTTPException(status_code=404, detail="No preview available")
            rel_path = rec["clip_filename"]
    path, reason = resolve_artifact_path(rel_path, "classic")
    if not path:
        raise HTTPException(status_code=404, detail="Classic video file missing (file_missing/stale_path/needs_regeneration)")
    return _stream_video_file(path, request)


@app.get("/api/groups/{group_id}/director-download")
async def download_director_video(group_id: int, request: Request):
    return await _download_artifact(group_id, "director", "director_final_video", "导演版", request)


@app.get("/api/groups/{group_id}/creative-download")
async def download_creative_video(group_id: int, request: Request):
    return await _download_artifact(group_id, "creative", "creative_final_video", "自编版", request)


@app.get("/api/groups/{group_id}/qianchuan-preview-download")
async def download_qianchuan_preview(group_id: int, request: Request):
    return await _download_artifact(group_id, "qianchuan_preview", "qianchuan_preview_video", "千川预览", request)


@app.get("/api/groups/{group_id}/qianchuan-download")
async def download_qianchuan_video(group_id: int, request: Request):
    return await _download_artifact(group_id, "qianchuan", "qianchuan_final_video", "千川结果", request)


@app.get("/api/groups/{group_id}/realistic-download")
async def download_realistic_video(group_id: int, request: Request):
    return await _download_artifact(group_id, "realistic", "realistic_final_video", "直出版", request)


@app.get("/api/groups/{group_id}/conservative-download")
async def download_conservative_video(group_id: int, request: Request):
    return await _download_artifact(group_id, "conservative", "conservative_final_video", "保守版", request)


async def _download_artifact(group_id: int, version: str, field: str, label: str, request: Request):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(f"SELECT {field} FROM clip_groups WHERE id = ?", (group_id,)) as cur:
            row = await cur.fetchone()
    path, reason = resolve_artifact_path(row[field] if row else None, version)
    if not path:
        if reason == "not_generated" and version == "qianchuan":
            detail = "千川结果尚未生成，请先生成/重试千川版。"
        elif reason == "not_generated":
            detail = f"{label}尚未生成，请先生成后重试。"
        else:
            detail = f"{label}文件缺失（file_missing/stale_path/needs_regeneration），请重新生成。"
        raise HTTPException(status_code=404, detail=detail)
    return _stream_video_file(path, request)


def _stream_video_file(path: str, request: Request):
    """Serve a verified video with browser-compatible byte-range streaming."""
    file_size = os.path.getsize(path)
    range_header = request.headers.get("range")
    start, end, status_code = 0, file_size - 1, 200
    if range_header:
        try:
            unit, value = range_header.split("=", 1)
            if unit.strip().lower() != "bytes" or "," in value:
                raise ValueError
            start_text, end_text = value.strip().split("-", 1)
            if start_text:
                start = int(start_text)
                end = int(end_text) if end_text else file_size - 1
            else:
                suffix_length = int(end_text)
                if suffix_length <= 0:
                    raise ValueError
                start = max(file_size - suffix_length, 0)
            if start < 0 or start >= file_size or end < start:
                raise ValueError
            end = min(end, file_size - 1)
            status_code = 206
        except (ValueError, TypeError):
            raise HTTPException(status_code=416, detail="Invalid Range header")
    content_length = end - start + 1
    headers = {
        "Accept-Ranges": "bytes",
        "Content-Length": str(content_length),
        "Content-Disposition": f'inline; filename="{os.path.basename(path)}"',
    }
    if status_code == 206:
        headers["Content-Range"] = f"bytes {start}-{end}/{file_size}"

    def iter_video():
        with open(path, "rb") as video_file:
            video_file.seek(start)
            remaining = content_length
            while remaining:
                chunk = video_file.read(min(1024 * 1024, remaining))
                if not chunk:
                    break
                remaining -= len(chunk)
                yield chunk

    return StreamingResponse(iter_video(), status_code=status_code, media_type="video/mp4", headers=headers)


@app.get("/api/recordings/processing-progress")
async def get_processing_progress():
    """Return progress for all currently processing recordings (transcribing + clipping)."""
    import time, statistics
    result = {}

    # ── Clipping progress (from in-memory dict) ──────────────────────────────
    for rid, p in _clip_progress.items():
        result[str(rid)] = {
            "phase": p.get("phase", ""),
            "pct": p.get("pct", 0),
            "msg": p.get("msg", ""),
            "eta_seconds": p.get("eta_seconds"),
        }

    # ── Transcription progress (time-based estimate) ──────────────────────────
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT id, gpu_job_id FROM recordings WHERE transcribed = 1 AND gpu_job_id IS NOT NULL"
        ) as cur:
            transcribing = await cur.fetchall()

    avg_dur = statistics.mean(_job_durations) if _job_durations else None
    now = time.time()
    for rec in transcribing:
        rid = rec["id"]
        job_id = rec["gpu_job_id"]
        if str(rid) in result:
            continue  # already has clip progress
        submit_at = _job_submit_times.get(job_id)
        if submit_at and avg_dur and avg_dur > 0:
            elapsed = now - submit_at
            pct = min(95, int(elapsed / avg_dur * 100))
            eta = max(0, int(avg_dur - elapsed))
        else:
            pct = 5
            eta = None
        result[str(rid)] = {
            "phase": "transcribe",
            "pct": pct,
            "msg": "GPU转录中",
            "eta_seconds": eta,
        }

    return result


@app.get("/api/recordings/{recording_id}")
async def get_recording(recording_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT r.*, rm.name as room_name FROM recordings r "
            "JOIN rooms rm ON r.room_id = rm.id WHERE r.id = ?",
            (recording_id,)
        ) as cur:
            rec = await cur.fetchone()
    if not rec:
        raise HTTPException(status_code=404, detail="Recording not found")
    return dict(rec)


# ── Retry ────────────────────────────────────────────────────────────────────

@app.post("/api/recordings/{recording_id}/retry-transcribe")
async def retry_transcribe(recording_id: int, regenerate_clip: bool = Query(default=True)):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM recordings WHERE id = ?", (recording_id,)) as cur:
            rec = await cur.fetchone()
    if not rec:
        raise HTTPException(status_code=404, detail="Recording not found")
    filepath = os.path.join(os.path.dirname(__file__), "..", "recordings", rec["filename"])
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Recording file missing on disk")
    # Preserve the old sidecar before the poll loop replaces it.
    srt_path = os.path.splitext(filepath)[0] + ".srt"
    backup_path = srt_path + ".previous.srt"
    if os.path.isfile(srt_path):
        try:
            shutil.copyfile(srt_path, backup_path)
        except OSError as exc:
            raise HTTPException(status_code=500, detail="Could not preserve existing SRT") from exc
    # Reset to unsynced/untranscribed so the poll loop picks it up via maybe_merge_before_upload
    # (which handles large-file splitting before GPU upload)
    async with aio_connect() as db:
        await db.execute(
            "UPDATE recordings SET transcribed = 0, synced = 0, gpu_job_id = NULL, clipped = 0 WHERE id = ?",
            (recording_id,)
        )
        await db.commit()
    if not regenerate_clip:
        from transcribe import _retranscribe_without_clip
        _retranscribe_without_clip.add(recording_id)
    from comfyui_client import free_vram
    from transcribe import flush_poll
    await free_vram()
    await flush_poll()   # wake the poll loop immediately
    return {"recording_id": recording_id, "status": "queued", "regenerate_clip": regenerate_clip}
    return {
        "recording_id": recording_id,
        "status": "queued",
        "regenerate_clip": regenerate_clip,
        "srt_backup": os.path.basename(backup_path) if os.path.isfile(backup_path) else None,
    }


@app.post("/api/recordings/{recording_id}/retranscribe")
async def retranscribe_recording(recording_id: int):
    """Queue a GPU-only ASR refresh and leave creative pipelines untouched.

    The existing SRT is retained as ``.previous.srt`` until the new GPU SRT
    is downloaded. This endpoint deliberately resets only this recording's
    transcription/clip state; it never changes the recording's group state or
    schedules director, creative, or Qianchuan work.
    """
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM recordings WHERE id = ?", (recording_id,)) as cur:
            rec = await cur.fetchone()
    if not rec:
        raise HTTPException(status_code=404, detail="Recording not found")

    recordings_dir = os.path.join(os.path.dirname(__file__), "..", "recordings")
    filepath = os.path.join(recordings_dir, rec["filename"])
    if not os.path.isfile(filepath):
        raise HTTPException(status_code=404, detail="Recording file missing on disk")

    srt_path = os.path.splitext(filepath)[0] + ".srt"
    backup_path = srt_path + ".previous.srt"
    if os.path.isfile(srt_path):
        try:
            shutil.copyfile(srt_path, backup_path)
        except OSError as exc:
            raise HTTPException(status_code=500, detail="Could not preserve existing SRT") from exc

    async with aio_connect() as db:
        await db.execute(
            """UPDATE recordings
               SET transcribed = 0, synced = 0, gpu_job_id = NULL,
                   clipped = 0, clip_filename = NULL, clip_error = NULL,
                   transcribe_error = NULL
               WHERE id = ?""",
            (recording_id,),
        )
        await db.commit()

    from comfyui_client import free_vram
    from transcribe import flush_poll
    await free_vram()
    await flush_poll()
    return {
        "recording_id": recording_id,
        "status": "queued",
        "srt_backup": os.path.basename(backup_path) if os.path.isfile(backup_path) else None,
        "creative_pipelines_triggered": False,
    }


@app.post("/api/recordings/clip-missing")
async def clip_missing():
    """查找所有已转录但未剪辑的记录，批量发起剪辑任务。"""
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        rows = await db.execute_fetchall(
            "SELECT * FROM recordings WHERE transcribed = 2 AND clipped IN (0, -1)"
        )
    queued, skipped = [], []
    base = os.path.join(os.path.dirname(__file__), "..", "recordings")
    for rec in rows:
        mp4_path = os.path.join(base, rec["filename"])
        srt_path = resolve_srt_path(os.path.join(base, rec["filename"]))
        if not os.path.exists(mp4_path) or not srt_path:
            skipped.append(rec["id"])
            continue
        clip_count = rec["clip_count"] if rec["clip_count"] else 1
        asyncio.create_task(_run_editor(rec["id"], mp4_path, srt_path, clip_count=clip_count, broadcast_fn=broadcast))
        queued.append(rec["id"])
    return {"queued": queued, "skipped": skipped}


@app.post("/api/recordings/{recording_id}/retry-clip")
async def retry_clip(recording_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM recordings WHERE id = ?", (recording_id,)) as cur:
            rec = await cur.fetchone()
    if not rec:
        raise HTTPException(status_code=404, detail="Recording not found")
    if rec["transcribed"] != 2:
        raise HTTPException(status_code=409, detail="Transcription not complete")
    mp4_path = os.path.join(os.path.dirname(__file__), "..", "recordings", rec["filename"])
    srt_path = resolve_srt_path(os.path.join(os.path.dirname(__file__), "..", "recordings", rec["filename"]))
    if not os.path.exists(mp4_path):
        raise HTTPException(status_code=404, detail="Recording file missing")
    if not srt_path:
        raise HTTPException(status_code=404, detail="SRT file missing")
    clip_count = rec["clip_count"] if rec["clip_count"] else 1
    asyncio.create_task(_run_editor(recording_id, mp4_path, srt_path, clip_count=clip_count, broadcast_fn=broadcast))
    return {"recording_id": recording_id, "clipped": 1}


@app.post("/api/recordings/{recording_id}/reveal-clip")
async def reveal_clip(recording_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM recordings WHERE id = ?", (recording_id,)) as cur:
            rec = await cur.fetchone()
    if not rec or rec["clipped"] != 2 or not rec["clip_filename"]:
        raise HTTPException(status_code=404, detail="Clip not found")
    clip_path = os.path.abspath(
        os.path.join(os.path.dirname(__file__), "..", "recordings", rec["clip_filename"])
    )
    if not os.path.exists(clip_path):
        raise HTTPException(status_code=404, detail="Clip file missing on disk")
    await asyncio.create_subprocess_exec("open", "-R", clip_path)
    return {"ok": True}


# ── Group management ──────────────────────────────────────────────────────────

class GroupCreate(BaseModel):
    room_id: int
    label: str
    wig_model: Optional[str] = None
    wig_color: Optional[str] = None


class GroupUpdate(BaseModel):
    label: str
    wig_model: Optional[str] = None
    wig_color: Optional[str] = None


class RecordingGroupUpdate(BaseModel):
    group_id: Optional[int] = None


class ImportVideosRequest(BaseModel):
    paths: list[str]


class CustomGroupCreate(BaseModel):
    label: str
    wig_model: Optional[str] = None
    wig_color: Optional[str] = None


async def _get_custom_room_id() -> int:
    """Get or create the special '自定义上传' room used for custom groups."""
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT id FROM rooms WHERE url = '__custom__'") as cur:
            row = await cur.fetchone()
        if row:
            return row["id"]
        cur = await db.execute(
            "INSERT INTO rooms (name, url, enabled) VALUES ('自定义上传', '__custom__', 0)",
        )
        await db.commit()
        return cur.lastrowid


@app.post("/api/groups/custom", status_code=201)
async def create_custom_group(body: CustomGroupCreate):
    room_id = await _get_custom_room_id()
    async with aio_connect() as db:
        cur = await db.execute(
            "INSERT INTO clip_groups (room_id, label, wig_model, wig_color, is_custom) VALUES (?, ?, ?, ?, 1)",
            (room_id, body.label, body.wig_model or None, body.wig_color or None),
        )
        await db.commit()
        return {"id": cur.lastrowid, "label": body.label,
                "wig_model": body.wig_model, "wig_color": body.wig_color, "is_custom": 1}


@app.post("/api/groups/{group_id}/upload-video", status_code=201)
async def upload_custom_group_video(group_id: int, file: UploadFile = File(...), clip_count: int = Form(1)):
    """Upload a video file directly to a custom group and trigger the clip pipeline."""
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM clip_groups WHERE id = ? AND is_custom = 1", (group_id,)) as cur:
            group = await cur.fetchone()
    if not group:
        raise HTTPException(status_code=404, detail="Custom group not found")

    room_id = group["room_id"]
    clip_count = max(1, min(5, clip_count))
    now = datetime.utcnow()
    ts = now.strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r"[^\w\-.]", "_", file.filename or "upload.mp4")
    filename = f"custom_{ts}_{safe_name}"
    recordings_dir = os.path.join(os.path.dirname(__file__), "..", "recordings")
    os.makedirs(recordings_dir, exist_ok=True)
    filepath = os.path.join(recordings_dir, filename)

    size_bytes = 0
    with open(filepath, "wb") as f:
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            f.write(chunk)
            size_bytes += len(chunk)

    start_time = now.isoformat()
    async with aio_connect() as db:
        cur = await db.execute(
            "INSERT INTO recordings (room_id, filename, start_time, end_time, size_bytes, synced, clip_count, group_id) VALUES (?, ?, ?, ?, ?, 0, ?, ?)",
            (room_id, filename, start_time, start_time, size_bytes, clip_count, group_id),
        )
        await db.commit()
        recording_id = cur.lastrowid

    asyncio.create_task(_generate_upload_thumb(recording_id, filepath))

    from comfyui_client import free_vram
    await free_vram()
    job_id = await sync_file(filepath, room_id)
    if job_id:
        async with aio_connect() as db:
            await db.execute(
                "UPDATE recordings SET transcribed = 1, synced = 1, gpu_job_id = ? WHERE id = ?",
                (job_id, recording_id),
            )
            await db.commit()

    return {"id": recording_id, "filename": filename, "size_bytes": size_bytes}


@app.post("/api/groups", status_code=201)
async def create_group(body: GroupCreate):
    async with aio_connect() as db:
        cur = await db.execute(
            "INSERT INTO clip_groups (room_id, label, wig_model, wig_color) VALUES (?, ?, ?, ?)",
            (body.room_id, body.label, body.wig_model or None, body.wig_color or None),
        )
        await db.commit()
        return {"id": cur.lastrowid, "label": body.label,
                "wig_model": body.wig_model, "wig_color": body.wig_color}


@app.post("/api/groups/{group_id}/import-videos", status_code=200)
async def import_group_videos(group_id: int, body: ImportVideosRequest):
    """Associate local .mp4 files with a group. Files outside recordings/ are copied in."""
    recordings_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "recordings"))
    imported = 0
    skipped = []
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT room_id FROM clip_groups WHERE id = ?", (group_id,)) as cur:
            grp = await cur.fetchone()
        if not grp:
            raise HTTPException(status_code=404, detail="Group not found")
        room_id = grp["room_id"]

        for raw_path in body.paths:
            path = raw_path.strip()
            if not path:
                continue
            if not os.path.isfile(path) or not path.lower().endswith(".mp4"):
                skipped.append(path)
                continue

            abs_path = os.path.abspath(path)
            # If outside recordings dir, copy it in
            if not abs_path.startswith(recordings_dir + os.sep):
                import shutil
                dest = os.path.join(recordings_dir, os.path.basename(abs_path))
                if not os.path.exists(dest):
                    shutil.copy2(abs_path, dest)
                abs_path = dest

            filename = os.path.basename(abs_path)
            size = os.path.getsize(abs_path)

            # Upsert: if already in DB, update group_id; else insert
            async with db.execute("SELECT id FROM recordings WHERE filename = ?", (filename,)) as cur:
                existing = await cur.fetchone()
            if existing:
                await db.execute(
                    "UPDATE recordings SET group_id = ?, local_deleted = 0 WHERE id = ?",
                    (group_id, existing["id"]),
                )
            else:
                imported_at = datetime.utcnow().isoformat()
                await db.execute(
                    """INSERT INTO recordings
                       (room_id, filename, start_time, end_time, size_bytes, group_id, synced, transcribed, clipped, local_deleted, segment_index)
                       VALUES (?, ?, ?, ?, ?, ?, 0, 0, 0, 0, 0)""",
                    (room_id, filename, imported_at, imported_at, size, group_id),
                )
            imported += 1

        await db.commit()
    return {"imported": imported, "skipped": skipped}


@app.patch("/api/groups/{group_id}")
async def update_group(group_id: int, body: GroupUpdate):
    async with aio_connect() as db:
        async with db.execute("SELECT id FROM clip_groups WHERE id = ?", (group_id,)) as cur:
            if not await cur.fetchone():
                raise HTTPException(status_code=404, detail="Group not found")
        await db.execute(
            "UPDATE clip_groups SET label = ?, wig_model = ?, wig_color = ? WHERE id = ?",
            (body.label, body.wig_model or None, body.wig_color or None, group_id),
        )
        await db.commit()
    return {"id": group_id, "label": body.label,
            "wig_model": body.wig_model, "wig_color": body.wig_color}


@app.delete("/api/groups/{group_id}", status_code=204)
async def delete_group(group_id: int):
    async with aio_connect() as db:
        async with db.execute("SELECT id FROM clip_groups WHERE id = ?", (group_id,)) as cur:
            if not await cur.fetchone():
                raise HTTPException(status_code=404, detail="Group not found")
        # Unlink recordings from this group (don't delete the recordings themselves)
        await db.execute("UPDATE recordings SET group_id = NULL WHERE group_id = ?", (group_id,))
        await db.execute("DELETE FROM clip_groups WHERE id = ?", (group_id,))
        await db.commit()


@app.delete("/api/recordings/{recording_id}/local-file", status_code=204)
async def delete_local_file(recording_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM recordings WHERE id = ?", (recording_id,)) as cur:
            rec = await cur.fetchone()
    if not rec:
        raise HTTPException(status_code=404, detail="Recording not found")
    if rec["synced"] != 1:
        raise HTTPException(status_code=409, detail="Not synced yet")
    if rec["transcribed"] == 1 or rec["clipped"] == 1:
        raise HTTPException(status_code=409, detail="Processing in progress")
    filepath = os.path.join(os.path.dirname(__file__), "..", "recordings", rec["filename"])
    if os.path.exists(filepath):
        os.unlink(filepath)
    async with aio_connect() as db:
        await db.execute("UPDATE recordings SET local_deleted = 1 WHERE id = ?", (recording_id,))
        await db.commit()


@app.post("/api/cleanup/local-files")
async def bulk_cleanup_local_files():
    """Delete local MP4s for recordings that are fully processed."""
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("""
            SELECT * FROM recordings
            WHERE synced = 1
              AND transcribed IN (2, -1)
              AND clipped IN (2, -1)
              AND local_deleted = 0
        """) as cur:
            candidates = await cur.fetchall()
    deleted = 0
    for rec in candidates:
        filepath = os.path.join(os.path.dirname(__file__), "..", "recordings", rec["filename"])
        if os.path.exists(filepath):
            os.unlink(filepath)
            deleted += 1
        async with aio_connect() as db:
            await db.execute("UPDATE recordings SET local_deleted = 1 WHERE id = ?", (rec["id"],))
            await db.commit()
    return {"deleted": deleted, "total_eligible": len(candidates)}


@app.patch("/api/recordings/{recording_id}/group")
async def reassign_recording_group(recording_id: int, body: RecordingGroupUpdate):
    async with aio_connect() as db:
        async with db.execute("SELECT id FROM recordings WHERE id = ?", (recording_id,)) as cur:
            if not await cur.fetchone():
                raise HTTPException(status_code=404, detail="Recording not found")
        await db.execute(
            "UPDATE recordings SET group_id = ? WHERE id = ?", (body.group_id, recording_id)
        )
        await db.commit()
    return {"recording_id": recording_id, "group_id": body.group_id}


# ── Reclip ───────────────────────────────────────────────────────────────────

class ReclipRequest(BaseModel):
    room_name: str
    date: str        # "YYYY-MM-DD"
    duration_sec: float
    clip_count: int = 1


@app.post("/api/reclip")
async def reclip(req: ReclipRequest):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT r.* FROM recordings r JOIN rooms rm ON r.room_id = rm.id "
            "WHERE rm.name = ? AND substr(r.start_time,1,10) = ? AND r.transcribed = 2",
            (req.room_name, req.date)
        ) as cur:
            recs = await cur.fetchall()
    if not recs:
        raise HTTPException(status_code=404, detail="No transcribed recordings found")
    queued = []
    for rec in recs:
        mp4_path = os.path.join(os.path.dirname(__file__), "..", "recordings", rec["filename"])
        srt_path = os.path.splitext(mp4_path)[0] + ".srt"
        if os.path.exists(mp4_path) and os.path.exists(srt_path):
            asyncio.create_task(_run_editor(rec["id"], mp4_path, srt_path, clip_duration=req.duration_sec, clip_count=req.clip_count, feedback=rec["reclip_feedback"], broadcast_fn=broadcast))
            queued.append(rec["id"])
    return {"queued": queued}


@app.post("/api/groups/{group_id}/reclip-all")
async def reclip_group_all(group_id: int):
    """Reset every clipped recording in a group back to pending and re-enqueue all."""
    recordings_dir = os.path.join(os.path.dirname(__file__), "..", "recordings")
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT id, filename, clip_count FROM recordings "
            "WHERE group_id = ? AND transcribed = 2",
            (group_id,),
        ) as cur:
            recs = await cur.fetchall()
        if not recs:
            raise HTTPException(status_code=404, detail="No transcribed recordings in group")
        ids = [r["id"] for r in recs]
        placeholders = ",".join("?" * len(ids))
        await db.execute(
            f"UPDATE recordings SET clipped=0, clip_filename=NULL, clip_error=NULL, skip_reason=NULL "
            f"WHERE id IN ({placeholders})",
            ids,
        )
        await db.execute(
            """UPDATE clip_groups SET
               merge_status=0, merged_filename=NULL, merged_at=NULL, merge_error=NULL,
               classic_status=0, director_status=0, director_error=NULL, director_final_video=NULL
               WHERE id=?""",
            (group_id,),
        )
        await db.commit()

    queued = []
    for rec in recs:
        mp4_path = os.path.join(recordings_dir, rec["filename"])
        srt_path = os.path.splitext(mp4_path)[0] + ".srt"
        if os.path.exists(mp4_path) and os.path.exists(srt_path):
            asyncio.create_task(_run_editor(
                rec["id"], mp4_path, srt_path,
                clip_count=rec["clip_count"] or 1,
                broadcast_fn=broadcast,
            ))
            queued.append(rec["id"])

    return {"group_id": group_id, "queued": queued, "total": len(recs)}


# ── GPU Status ────────────────────────────────────────────────────────────────

GPU_SERVICE_URL = os.environ.get("GPU_SERVICE_URL", "http://10.190.0.203:8877")


COMFYUI_URL = os.environ.get("COMFYUI_URL", "http://10.190.0.203:8188")


@app.get("/api/gpu/status")
async def gpu_status():
    from gpu_state import is_online as gpu_is_online, is_maintenance as gpu_is_maint, _offline_since
    from gpu_state import WATCHDOG_URL
    import time as _time
    offline_sec = int(_time.monotonic() - _offline_since) if (not gpu_is_online() and _offline_since) else 0
    result = {
        "reachable": False, "health": {}, "jobs": [],
        "gpu_online": gpu_is_online(),
        "online": False,
        "gpu_service_url": GPU_SERVICE_URL,
        "gpu_offline_seconds": offline_sec,
        "deployment_role": _DEPLOYMENT_ROLE,
        "remote_backend_url": REMOTE_BACKEND_URL,
        "comfyui": {"reachable": False, "vram_total": 0, "vram_free": 0, "ram_total": 0, "ram_free": 0, "queue_running": 0, "queue_pending": 0},
    }
    # Probe the remote service on every request; watcher hysteresis is for
    # scheduling, not for presenting stale status to operators.
    skip_gpu_probe = gpu_is_maint()
    try:
        # Use the established async HTTP client dependency rather than
        # aiohttp.  aiohttp can be absent/misconfigured on the control-plane
        # host, which previously made healthy remote ComfyUI endpoints look
        # unreachable.  Keep this probe remote-only; there is no local fallback.
        async def _aio_get(url):
            try:
                async with httpx.AsyncClient(timeout=5) as _s:
                    _r = await _s.get(url)
                    try:
                        _body = _r.json()
                    except ValueError:
                        _body = None
                    return _r.status_code, _body
            except Exception as _e:
                return None, _e

        async def _skip(): return (None, None)
        comfy_r, queue_r = await asyncio.gather(
            _aio_get(f"{COMFYUI_URL}/system_stats") if not gpu_is_maint() else _skip(),
            _aio_get(f"{COMFYUI_URL}/queue") if not gpu_is_maint() else _skip(),
            return_exceptions=True,
        )
        # Probe port 8877 on every request so status is not stale during the
        # watcher startup/hysteresis window or after a watchdog restart. This
        # only observes the remote worker; failures report it offline and never
        # enable local processing or fallback.
        try:
            _st, _body = await _aio_get(f"{GPU_SERVICE_URL}/health")
            if _st == 200:
                result["reachable"] = True
                result["online"] = True
                result["gpu_online"] = True
                result["health"] = _body if isinstance(_body, dict) else {}
        except Exception:
            pass

        # Unpack aio_get tuples
        comfy_status, comfy_body = comfy_r if isinstance(comfy_r, tuple) else (None, None)
        queue_status, queue_body = queue_r if isinstance(queue_r, tuple) else (None, None)
        if comfy_status == 200 and isinstance(comfy_body, dict):
            cs = comfy_body
            dev = cs.get("devices", [{}])[0]
            sys_ = cs.get("system", {})
            # Use torch_vram (dedicated/allocated by PyTorch) rather than the
            # shared-memory pool that AMD iGPU reports as vram_total/vram_free.
            # Hardware total is always reliable; use it as the denominator.
            # torch_vram fields reflect only PyTorch-allocated memory (0 when
            # no models are loaded), so don't use them as the "total" baseline.
            vram_total = dev.get("vram_total", 0)
            vram_free  = dev.get("vram_free",  0)
            # If the hardware fields are missing (some virtual devices), fall
            # back to torch allocation fields.
            if not vram_total:
                vram_total = dev.get("torch_vram_total", 0)
                vram_free  = dev.get("torch_vram_free",  0)
            result["comfyui"] = {
                "reachable": True,
                "vram_total": vram_total,
                "vram_free": vram_free,
                "ram_total": sys_.get("ram_total", 0),
                "ram_free": sys_.get("ram_free", 0),
                "queue_running": 0,
                "queue_pending": 0,
            }
        if queue_status == 200 and isinstance(queue_body, dict):
            # /queue is also a valid liveness signal.  Do not require
            # /system_stats (which may be temporarily unavailable while
            # ComfyUI remains healthy) to report the service as reachable.
            result["comfyui"]["reachable"] = True
            result["comfyui"]["queue_running"] = len(queue_body.get("queue_running", []))
            result["comfyui"]["queue_pending"] = len(queue_body.get("queue_pending", []))

        # Keep the watchdog's view alongside the direct probes for diagnosing
        # split-brain/network-path failures. It is diagnostic only and never
        # enables local processing or replaces the remote GPU requirement.
        watchdog_status_code, watchdog_body = await _aio_get(f"{WATCHDOG_URL}/status")
        result["watchdog_probe"] = {
            "reachable": watchdog_status_code == 200,
            "services": watchdog_body if watchdog_status_code == 200 and isinstance(watchdog_body, dict) else {},
        }
        comfy_watchdog = (
            watchdog_body.get("comfyui")
            if watchdog_status_code == 200 and isinstance(watchdog_body, dict)
            else None
        )
        if isinstance(comfy_watchdog, dict):
            result["comfyui"]["watchdog_healthy"] = bool(comfy_watchdog.get("healthy"))
    except Exception as e:
        logger.error(f"GPU status check failed: {e}")

    # Pending transcription jobs: in-flight on GPU + waiting to upload (exclude live segments)
    async with aio_connect() as db:
        async with db.execute(
            """SELECT COUNT(*) FROM recordings
               WHERE (transcribed = 1 AND gpu_job_id IS NOT NULL)
                  OR (transcribed = 0 AND synced = 0 AND local_deleted = 0 AND end_time IS NOT NULL)"""
        ) as cur:
            (pending_transcribe,) = await cur.fetchone()
    result["pending_transcribe"] = pending_transcribe
    # Include cached watchdog state (no extra HTTP call needed)
    from gpu_state import watchdog_status
    result["watchdog"] = watchdog_status()
    # Include poll loop health state
    import time as _t
    from datetime import datetime, timezone
    now = _t.time()
    ps = _poll_state
    def _iso(ts):
        return datetime.fromtimestamp(ts, tz=timezone.utc).isoformat() if ts else None
    result["poll_state"] = {
        "last_poll_at": _iso(ps["last_poll_at"]),
        "last_submit_at": _iso(ps["last_submit_at"]),
        "last_complete_at": _iso(ps["last_complete_at"]),
        "blocked_count": ps["blocked_count"],
        "active_job_id": ps["active_job_id"],
        "poll_interval": POLL_INTERVAL,
    }
    # Maintenance mode flag
    from gpu_state import is_maintenance
    result["maintenance"] = is_maintenance()
    return result


@app.post("/api/gpu/maintenance")
async def gpu_maintenance_enable():
    """Enable GPU maintenance mode: skips all GPU pipelines and silences offline alerts."""
    from gpu_state import set_maintenance
    set_maintenance(True)
    logger.info("GPU maintenance mode enabled via API")
    return {"ok": True, "maintenance": True}


@app.delete("/api/gpu/maintenance")
async def gpu_maintenance_disable():
    """Disable GPU maintenance mode: resume normal GPU operation."""
    from gpu_state import set_maintenance
    set_maintenance(False)
    logger.info("GPU maintenance mode disabled via API")
    return {"ok": True, "maintenance": False}


@app.post("/api/transcribe/flush", status_code=200)
async def flush_transcribe_queue():
    """Wake the poll loop immediately without waiting for the 60-second interval."""
    await flush_poll()
    return {"ok": True, "msg": "Poll loop woken up"}


@app.get("/api/watchdog/status")
async def get_watchdog_status():
    """Proxy to watchdog agent /status."""
    from gpu_state import WATCHDOG_URL, watchdog_status
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            r = await client.get(f"{WATCHDOG_URL}/status")
            if r.status_code == 200:
                return r.json()
    except Exception:
        pass
    return watchdog_status().get("services", {})


@app.get("/api/watchdog/ping")
async def watchdog_ping():
    """Check remote watchdog availability without turning an outage into a 500."""
    from gpu_state import WATCHDOG_URL, watchdog_status
    from watchdog import ping_watchdog
    result = await ping_watchdog(WATCHDOG_URL)
    cached = watchdog_status()
    result["cached"] = cached
    result["available"] = result["reachable"]
    return result


@app.post("/api/watchdog/start/{service}")
async def watchdog_start(service: str):
    """Ask watchdog agent to start a named service."""
    from gpu_state import WATCHDOG_URL
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            r = await client.post(f"{WATCHDOG_URL}/start/{service}")
            if r.status_code == 200:
                return r.json()
            raise HTTPException(status_code=r.status_code, detail=r.text)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Watchdog unreachable: {e}")


@app.post("/api/watchdog/stop/{service}")
async def watchdog_stop(service: str):
    """Ask watchdog agent to stop a named service."""
    from gpu_state import WATCHDOG_URL
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            r = await client.post(f"{WATCHDOG_URL}/stop/{service}")
            if r.status_code == 200:
                return r.json()
            raise HTTPException(status_code=r.status_code, detail=r.text)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Watchdog unreachable: {e}")


@app.post("/api/watchdog/restart/{service}")
async def watchdog_restart(service: str):
    """Ask watchdog agent to restart a named service."""
    from gpu_state import WATCHDOG_URL
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            r = await client.post(f"{WATCHDOG_URL}/restart/{service}")
            if r.status_code == 200:
                return r.json()
            raise HTTPException(status_code=r.status_code, detail=r.text)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Watchdog unreachable: {e}")


@app.get("/api/gpu/logs")
async def gpu_logs():
    """Recent transcription activity for the GPU log marquee."""
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            """SELECT r.id, r.filename, r.transcribed, r.transcribe_error, r.start_time, rm.name as room_name
               FROM recordings r LEFT JOIN rooms rm ON r.room_id = rm.id
               WHERE r.gpu_job_id IS NOT NULL
               ORDER BY r.id DESC LIMIT 20"""
        ) as cur:
            rows = await cur.fetchall()

    logs = []
    for row in rows:
        if row["transcribed"] == 2:
            status = "转录完成"
            level = "success"
        elif row["transcribed"] == -1:
            err = (row["transcribe_error"] or "")[:60]
            status = f"转录失败: {err}"
            level = "error"
        elif row["transcribed"] == 1:
            status = "转录中"
            level = "info"
        else:
            status = "等待转录"
            level = "pending"
        logs.append({
            "id": row["id"],
            "filename": row["filename"],
            "room": row["room_name"] or "",
            "status": status,
            "level": level,
            "time": (row["start_time"] or "")[:16].replace("T", " "),
        })
    return logs


# ── Version ──────────────────────────────────────────────────────────────────

@app.get("/api/version")
def get_version():
    return {"version": APP_VERSION}


# ── Status ───────────────────────────────────────────────────────────────────

@app.get("/api/status")
async def system_status():
    async with aio_connect() as db:
        async with db.execute("SELECT COUNT(*) FROM rooms WHERE enabled = 1") as cur:
            (enabled_rooms,) = await cur.fetchone()
        async with db.execute("SELECT COUNT(*) FROM recordings") as cur:
            (total_recordings,) = await cur.fetchone()
        async with db.execute("SELECT SUM(size_bytes) FROM recordings WHERE size_bytes IS NOT NULL") as cur:
            (total_bytes,) = await cur.fetchone()

    recordings_dir = os.path.join(os.path.dirname(__file__), "..", "recordings")
    local_files = len(os.listdir(recordings_dir)) if os.path.exists(recordings_dir) else 0

    return {
        "enabled_rooms": enabled_rooms,
        "active_recordings": sum(1 for rid in monitor._recorders if monitor._recorders[rid].recording),
        "total_recordings": total_recordings,
        "total_bytes": total_bytes or 0,
        "local_files": local_files,
    }


@app.get("/api/memory/status")
async def memory_status():
    """Return current system memory usage and clip dispatch pressure state."""
    try:
        import psutil
        vm = psutil.virtual_memory()
        used_gb  = round(vm.used  / 1024 ** 3, 2)
        total_gb = round(vm.total / 1024 ** 3, 2)
        avail_gb = round(vm.available / 1024 ** 3, 2)
    except ImportError:
        used_gb = total_gb = avail_gb = None

    from transcribe import _memory_pressure
    return {
        "used_gb": used_gb,
        "total_gb": total_gb,
        "available_gb": avail_gb,
        "pressure": _memory_pressure,
        "warn_threshold_gb": 20,
        "recover_threshold_gb": 17,
    }


@app.get("/api/clip-jobs")
async def get_clip_jobs():
    """Return in-progress clip job progress keyed by recording_id."""
    return _clip_progress


@app.get("/api/transcribe-queue")
async def get_transcribe_queue(limit: int = Query(default=100, ge=1, le=500)):
    """Return pending/running transcription jobs for the queue view."""
    import time as _time
    from transcribe import transcribe_timing
    limit = limit if isinstance(limit, int) else 100
    timing = transcribe_timing()
    now = _time.time()

    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            """SELECT r.id, r.filename, r.transcribed, r.transcribe_error, r.start_time,
                      r.end_time, r.gpu_job_id, r.synced, r.local_deleted, r.size_bytes,
                      rm.name as room_name
               FROM recordings r LEFT JOIN rooms rm ON r.room_id = rm.id
               WHERE (r.transcribed IN (0, 1) AND (r.synced = 1 OR r.gpu_job_id IS NOT NULL))
                  OR (r.transcribed = 0
                      AND r.synced = 0
                      AND r.local_deleted = 0
                      AND r.end_time IS NOT NULL
                      AND r.end_time != r.start_time)
               ORDER BY r.id ASC
               LIMIT ?""", (limit,)
        ) as cur:
            rows = await cur.fetchall()

    avg_s = timing["avg_duration_s"]
    submit_times = timing["submit_times"]

    jobs = []
    queue_pos = 0  # position among waiting-for-GPU jobs
    for row in rows:
        if row["synced"] == 0 and row["transcribed"] == 0 and not _is_finished_unsynced_upload_candidate(row):
            logger.info(
                "Transcribe queue skipped non-processable recording %s (%s): "
                "not finished, zero-duration, deleted, or missing local MP4",
                row["id"],
                row["filename"],
            )
            continue

        if row["transcribed"] == 1 and row["gpu_job_id"]:
            status = "转录中"
            level = "running"
            elapsed_s = int(now - submit_times[row["gpu_job_id"]]) if row["gpu_job_id"] in submit_times else None
            pct = min(99, int(elapsed_s / avg_s * 100)) if (elapsed_s is not None and avg_s > 0) else None
            pos = None
        elif row["transcribed"] == 0 and row["synced"] == 1:
            status = "等待转录"
            level = "queued"
            elapsed_s = None
            pct = None
            queue_pos += 1
            pos = queue_pos
        else:
            status = "待上传"
            level = "pending"
            elapsed_s = None
            pct = None
            queue_pos += 1
            pos = queue_pos

        jobs.append({
            "recording_id": row["id"],
            "filename": row["filename"],
            "room_name": row["room_name"] or "",
            "status": status,
            "level": level,
            "gpu_job_id": row["gpu_job_id"],
            "start_time": (row["start_time"] or "")[:16].replace("T", " "),
            "elapsed_s": elapsed_s,
            "pct": pct,
            "queue_pos": pos,
            "size_bytes": row["size_bytes"],
        })

    waiting_jobs = sum(1 for job in jobs if job["level"] != "running")
    running_remaining = sum(
        max(0.0, avg_s - (job["elapsed_s"] or 0))
        for job in jobs if job["level"] == "running"
    )
    eta_s = int(running_remaining + avg_s * waiting_jobs) if avg_s > 0 else None
    total = len(jobs) + timing["session_done"]
    return {
        "jobs": jobs,
        "avg_duration_s": avg_s,
        "session_done": timing["session_done"],
        "total": total,
        "eta_seconds": eta_s,
    }


@app.get("/api/clip-queue")
async def get_clip_queue_api(limit: int = Query(default=200, ge=1, le=500)):
    """Return running + queued clip jobs with priority info."""
    limit = limit if isinstance(limit, int) else 200
    queue = get_clip_queue()
    queue["queued"] = queue["queued"][:limit]
    queue["paused"] = queue["paused"][:limit]
    queue["total_queued"] = len(queue["queued"]) + len(queue["paused"])
    return queue


@app.post("/api/clip-queue/{recording_id}/priority")
async def set_clip_priority(recording_id: int, priority: int):
    """Update the priority of a queued clip job (1=highest, 99=lowest)."""
    priority = max(1, min(99, priority))
    updated = await update_job_priority(recording_id, priority)
    if not updated:
        raise HTTPException(status_code=404, detail="Job not found in queue (may already be running or completed)")
    return {"ok": True, "recording_id": recording_id, "priority": priority}


@app.post("/api/clip-queue/{recording_id}/cancel")
async def cancel_clip_queue_job(recording_id: int):
    """Remove a queued/paused job from the clip queue (cannot cancel running jobs)."""
    removed = await cancel_clip_job(recording_id)
    if not removed:
        raise HTTPException(status_code=404, detail="Job not found in queue (may be running or already completed)")
    # Reset clipped status so the job can be re-dispatched later
    async with aio_connect() as db:
        await db.execute("UPDATE recordings SET clipped = 0 WHERE id = ? AND clipped = 1", (recording_id,))
        await db.commit()
    return {"ok": True, "recording_id": recording_id}


@app.post("/api/clip-queue/{recording_id}/pause")
async def pause_clip_queue_job(recording_id: int):
    """Pause a queued job (it stays in queue but won't be dispatched until resumed)."""
    paused = await pause_clip_job(recording_id)
    if not paused:
        raise HTTPException(status_code=404, detail="Job not found in queue or is already running")
    return {"ok": True, "recording_id": recording_id, "status": "paused"}


@app.post("/api/clip-queue/{recording_id}/start")
async def start_clip_queue_job(recording_id: int):
    """Move a queued/paused job to the front of the queue (priority=1) and resume if paused."""
    # Resume if paused
    await resume_clip_job(recording_id)
    # Set to highest priority
    updated = await update_job_priority(recording_id, 1)
    if not updated:
        raise HTTPException(status_code=404, detail="Job not found in queue")
    return {"ok": True, "recording_id": recording_id, "priority": 1}


@app.post("/api/clip-queue/retry-all")
async def retry_all_clip_queue_jobs():
    """Re-enqueue failed clip jobs that have no skip reason and still have source files."""
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT id, filename, clip_count FROM recordings "
            "WHERE clipped = -1 AND (skip_reason IS NULL OR skip_reason = '')"
        ) as cur:
            records = await cur.fetchall()

    queued = 0
    failed = 0
    for rec in records:
        mp4_path = os.path.join(RECORDINGS_DIR, rec["filename"])
        srt_path = os.path.join(RECORDINGS_DIR, os.path.splitext(rec["filename"])[0] + ".srt")
        if not os.path.exists(mp4_path) or not os.path.exists(srt_path):
            failed += 1
            continue
        async with aio_connect() as db:
            await db.execute(
                "UPDATE recordings SET clipped = 0, clip_error = NULL WHERE id = ?",
                (rec["id"],),
            )
            await db.commit()
        asyncio.create_task(_run_editor(
            rec["id"], mp4_path, srt_path, clip_count=rec["clip_count"] or 1
        ))
        queued += 1
    return {"ok": True, "queued": queued, "failed": failed}


@app.post("/api/clip-queue/dismiss-all")
async def dismiss_all_clip_queue_jobs():
    """Mark all visible failed clip jobs as manually cleared."""
    async with aio_connect() as db:
        cursor = await db.execute(
            "UPDATE recordings SET clipped = -1, skip_reason = '已手动清除' "
            "WHERE clipped = -1 AND (skip_reason IS NULL OR skip_reason != '已手动清除')"
        )
        await db.commit()
        cleared = cursor.rowcount
    return {"ok": True, "cleared": cleared}


@app.post("/api/clip-queue/{recording_id}/retry")
async def retry_clip_queue_job(recording_id: int):
    """Re-enqueue a failed clip job (clipped=-1)."""
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT r.*, rm.name as room_name FROM recordings r "
            "JOIN rooms rm ON r.room_id = rm.id WHERE r.id = ?",
            (recording_id,)
        ) as cur:
            rec = await cur.fetchone()
    if not rec:
        raise HTTPException(status_code=404, detail="Recording not found")
    if rec["clipped"] != -1:
        raise HTTPException(status_code=409, detail=f"Recording is not in failed state (clipped={rec['clipped']})")

    mp4_path = os.path.join(RECORDINGS_DIR, rec["filename"])
    srt_filename = os.path.splitext(rec["filename"])[0] + ".srt"
    srt_path = os.path.join(RECORDINGS_DIR, srt_filename)
    if not os.path.exists(mp4_path):
        raise HTTPException(status_code=404, detail="MP4 file not found on disk")
    if not os.path.exists(srt_path):
        raise HTTPException(status_code=404, detail="SRT file not found — re-transcribe first")

    # Reset failed state and re-enqueue
    async with aio_connect() as db:
        await db.execute("UPDATE recordings SET clipped = 0, clip_error = NULL WHERE id = ?", (recording_id,))
        await db.commit()

    asyncio.create_task(_run_editor(recording_id, mp4_path, srt_path,
                                    clip_count=rec["clip_count"] or 1))
    return {"ok": True, "recording_id": recording_id, "status": "queued"}


async def _parse_bulk_clip_ids(body: dict) -> list[int]:
    """Validate the bounded recording-id list used by bulk clip actions."""
    ids = body.get("ids") if isinstance(body, dict) else None
    if not isinstance(ids, list) or not ids or len(ids) > 500:
        raise HTTPException(status_code=400, detail="ids must contain 1 to 500 recording IDs")
    if any(isinstance(recording_id, bool) or not isinstance(recording_id, int) or recording_id < 1 for recording_id in ids):
        raise HTTPException(status_code=400, detail="ids must contain positive integers")
    return list(dict.fromkeys(ids))


@app.post("/api/clip-queue/bulk-retry")
async def bulk_retry_clip_queue_jobs(body: dict = Body(...)):
    """Re-enqueue the eligible failed clip jobs supplied by the queue page."""
    recording_ids = await _parse_bulk_clip_ids(body)
    queued, skipped = [], []
    base = RECORDINGS_DIR
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        placeholders = ",".join("?" for _ in recording_ids)
        async with db.execute(
            f"SELECT * FROM recordings WHERE id IN ({placeholders}) AND clipped = -1",
            recording_ids,
        ) as cur:
            records = {row["id"]: row for row in await cur.fetchall()}

    for recording_id in recording_ids:
        rec = records.get(recording_id)
        if not rec or rec["skip_reason"]:
            skipped.append(recording_id)
            continue
        mp4_path = os.path.join(base, rec["filename"])
        srt_path = resolve_srt_path(os.path.join(base, rec["filename"]))
        if not os.path.exists(mp4_path) or not srt_path:
            skipped.append(recording_id)
            continue
        async with aio_connect() as db:
            await db.execute("UPDATE recordings SET clipped = 0, clip_error = NULL WHERE id = ?", (recording_id,))
            await db.commit()
        asyncio.create_task(_run_editor(recording_id, mp4_path, srt_path, clip_count=rec["clip_count"] or 1))
        queued.append(recording_id)
    return {"ok": True, "queued": queued, "skipped": skipped}


@app.post("/api/clip-queue/{recording_id}/dismiss")
async def dismiss_clip_job(recording_id: int):
    """Permanently dismiss a failed clip job so it no longer appears in the failed list."""
    async with aio_connect() as db:
        async with db.execute("SELECT clipped FROM recordings WHERE id = ?", (recording_id,)) as cur:
            rec = await cur.fetchone()
        if not rec:
            raise HTTPException(status_code=404, detail="Recording not found")
        await db.execute(
            "UPDATE recordings SET clipped = -1, skip_reason = '已手动清除' WHERE id = ?",
            (recording_id,),
        )
        await db.commit()
    return {"ok": True, "recording_id": recording_id}


@app.post("/api/clip-queue/bulk-dismiss")
async def bulk_dismiss_clip_jobs(body: dict = Body(...)):
    """Hide the failed clip jobs supplied by the queue page in one transaction."""
    recording_ids = await _parse_bulk_clip_ids(body)
    placeholders = ",".join("?" for _ in recording_ids)
    async with aio_connect() as db:
        async with db.execute(
            f"SELECT id FROM recordings WHERE id IN ({placeholders}) AND clipped = -1",
            recording_ids,
        ) as cur:
            dismissed = [row[0] for row in await cur.fetchall()]
        if dismissed:
            dismissed_placeholders = ",".join("?" for _ in dismissed)
            await db.execute(
                f"UPDATE recordings SET skip_reason = '已手动清除' WHERE id IN ({dismissed_placeholders})",
                dismissed,
            )
            await db.commit()
    return {"ok": True, "dismissed": dismissed}


# ── 画质增强 ──────────────────────────────────────────────────────────────────

ENHANCE_OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "enhance_output")

# 内存中跟踪作业状态（local_job_id → metadata）
_enhance_jobs: dict = {}
# 串行队列：每次只向 GPU 提交一个增强作业，避免 CUDA OOM
_enhance_queue: asyncio.Queue = asyncio.Queue()
_enhance_seq: int = 0


def _refresh_enhance_queue_positions():
    """重新计算所有排队中作业的位置编号。"""
    pos = 1
    for job in _enhance_jobs.values():
        if job.get("status") == "queued":
            job["queue_pos"] = pos
            job["msg"] = f"排队中，第 {pos} 位"
            pos += 1


async def _enhance_worker():
    """串行消费 enhance 队列，一次只处理一个作业，防止 GPU CUDA OOM。"""
    from enhance import submit_enhance_job, get_enhance_job_status, download_enhance_result
    while True:
        local_id, file_path, out_path = await _enhance_queue.get()
        job = _enhance_jobs.get(local_id)
        if not job or job.get("status") == "cancelled":
            if file_path and os.path.exists(file_path):
                os.unlink(file_path)
            _refresh_enhance_queue_positions()
            continue

        try:
            job.update({"status": "uploading", "msg": "上传中…", "queue_pos": 0, "pct": 0})
            _refresh_enhance_queue_positions()

            gpu_job_id = await submit_enhance_job(
                file_path,
                model=job["model"], target_res=job["target_res"],
                denoise=job["denoise"], preview_only=job["preview_only"],
            )
            if file_path and os.path.exists(file_path):
                os.unlink(file_path)

            if not gpu_job_id:
                job.update({"status": "error", "error": "提交到 GPU 增强服务失败"})
                continue

            job.update({"gpu_job_id": gpu_job_id, "status": "running", "msg": "GPU 处理中…"})

            # 轮询直到完成
            deadline = _time.time() + 7200
            consecutive_failures = 0
            while _time.time() < deadline:
                await asyncio.sleep(5)
                if job.get("status") == "cancelled":
                    from enhance import delete_enhance_job
                    await delete_enhance_job(gpu_job_id)
                    break
                data = await get_enhance_job_status(gpu_job_id)
                if not data:
                    consecutive_failures += 1
                    if consecutive_failures >= 24:
                        job.update({"status": "error", "error": "服务无响应超过 2 分钟，作业已中止"})
                        logger.warning(f"Enhance {local_id} aborted: unreachable 2 min")
                    continue
                consecutive_failures = 0
                status = data.get("status")
                job.update({"gpu_status": status, "pct": data.get("pct", 0), "msg": data.get("msg", "")})
                if status == "done":
                    ok = await download_enhance_result(gpu_job_id, out_path)
                    if ok:
                        job.update({"status": "done", "local_path": out_path})
                        logger.info(f"Enhance {local_id} done → {out_path}")
                    else:
                        job.update({"status": "error", "error": "下载结果失败"})
                    break
                if status == "error":
                    job.update({"status": "error", "error": data.get("error", "GPU 处理失败")})
                    break
            else:
                job.update({"status": "error", "error": "轮询超时"})

        except Exception as e:
            job.update({"status": "error", "error": str(e)})
            logger.error(f"Enhance worker error {local_id}: {e}")
        finally:
            _refresh_enhance_queue_positions()


@app.post("/api/enhance-jobs", status_code=201)
async def create_enhance_job(
    file:         UploadFile = File(...),
    model:        str = Form("general"),
    target_res:   str = Form("1080p"),
    denoise:      str = Form("medium"),
    preview_only: bool = Form(False),
):
    """接收文件，加入本地串行队列，返回 job_id 供前端轮询。"""
    global _enhance_seq
    from enhance import is_enhance_service_available
    if not await is_enhance_service_available():
        raise HTTPException(status_code=503, detail="画质增强服务不可用，请确认 GPU 服务器已启动 enhance_service.py")

    os.makedirs(ENHANCE_OUTPUT_DIR, exist_ok=True)
    safe_name = re.sub(r"[^\w\-.]", "_", file.filename or "upload")
    _enhance_seq += 1
    local_id  = f"enhance_{_enhance_seq}_{safe_name}"
    tmp_path  = os.path.join(ENHANCE_OUTPUT_DIR, f"_tmp_{local_id}")

    with open(tmp_path, "wb") as f:
        while chunk := await file.read(1024 * 1024):
            f.write(chunk)

    ext      = os.path.splitext(safe_name)[1]
    suffix   = "_preview" if preview_only else "_enhanced"
    out_name = os.path.splitext(safe_name)[0] + suffix + ext
    out_path = os.path.join(ENHANCE_OUTPUT_DIR, out_name)

    queue_pos = _enhance_queue.qsize() + 1
    _enhance_jobs[local_id] = {
        "job_id":       local_id,
        "status":       "queued",
        "queue_pos":    queue_pos,
        "gpu_job_id":   None,
        "gpu_status":   None,
        "pct":          0,
        "msg":          f"排队中，第 {queue_pos} 位",
        "filename":     safe_name,
        "out_filename": out_name,
        "model":        model,
        "target_res":   target_res,
        "denoise":      denoise,
        "preview_only": preview_only,
        "local_path":   None,
        "error":        None,
    }
    await _enhance_queue.put((local_id, tmp_path, out_path))
    logger.info(f"Enhance job queued: {local_id} (queue_pos={queue_pos})")
    return {"job_id": local_id, "filename": safe_name}


@app.get("/api/enhance-jobs/{job_id}")
async def get_enhance_job(job_id: str):
    job = _enhance_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job


@app.get("/api/enhance-jobs/{job_id}/download")
async def download_enhance_job(job_id: str):
    from fastapi.responses import FileResponse
    job = _enhance_jobs.get(job_id)
    if not job or job.get("status") != "done":
        raise HTTPException(status_code=404, detail="结果尚未就绪")
    path = job.get("local_path")
    if not path or not os.path.exists(path):
        raise HTTPException(status_code=404, detail="本地文件不存在")
    return FileResponse(path, filename=job["out_filename"])


@app.delete("/api/enhance-jobs/{job_id}")
async def cancel_enhance_job(job_id: str):
    job = _enhance_jobs.get(job_id)
    if job:
        gpu_id = job.get("gpu_job_id")
        if gpu_id:
            from enhance import delete_enhance_job
            await delete_enhance_job(gpu_id)
        job["status"] = "cancelled"   # worker 检查此标志跳过
        _enhance_jobs.pop(job_id, None)
    _refresh_enhance_queue_positions()
    return {"ok": True}


@app.get("/api/enhance-service/status")
async def enhance_service_status():
    from enhance import is_enhance_service_available, ENHANCE_SERVICE_URL
    available = await is_enhance_service_available()
    return {"available": available, "url": ENHANCE_SERVICE_URL}


@app.get("/api/stats")
async def get_stats():
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        rows = await db.execute_fetchall("""
            SELECT
                SUM(CASE WHEN transcribed = 0 AND local_deleted = 0 AND end_time IS NOT NULL THEN 1 ELSE 0 END) AS transcribe_pending,
                SUM(CASE WHEN transcribed = 1 THEN 1 ELSE 0 END)               AS transcribe_running,
                SUM(CASE WHEN transcribed = -1 THEN 1 ELSE 0 END)              AS transcribe_failed,
                SUM(CASE WHEN transcribed = 2 AND clipped = 0 THEN 1 ELSE 0 END) AS clip_pending,
                SUM(CASE WHEN clipped = 1 THEN 1 ELSE 0 END)                   AS clip_running,
                SUM(CASE WHEN clipped = -1 THEN 1 ELSE 0 END)                  AS clip_failed
            FROM recordings
        """)
        r = dict(rows[0])
    return {k: (v or 0) for k, v in r.items()}


# ── WebSocket ────────────────────────────────────────────────────────────────

@app.websocket("/ws/events")
async def ws_events(websocket: WebSocket):
    await websocket.accept()
    _ws_clients.add(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        _ws_clients.discard(websocket)


# ── Products ─────────────────────────────────────────────────────────────────

@app.get("/api/products")
async def list_products(keyword: Optional[str] = None):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        base = """SELECT p.*, rm.name as room_name
                  FROM products p
                  LEFT JOIN rooms rm ON p.room_id = rm.id"""
        if keyword:
            async with db.execute(
                base + " WHERE p.product_name LIKE ? OR p.keywords LIKE ? ORDER BY p.created_at DESC",
                (f"%{keyword}%", f"%{keyword}%"),
            ) as cur:
                rows = await cur.fetchall()
        else:
            async with db.execute(base + " ORDER BY p.created_at DESC") as cur:
                rows = await cur.fetchall()
    return [dict(r) for r in rows]


@app.post("/api/products", status_code=201)
async def create_product(body: ProductCreate):
    async with aio_connect() as db:
        cur = await db.execute(
            """INSERT INTO products (platform, product_id, product_name, product_url, product_thumb, keywords, enabled, room_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (body.platform, body.product_id, body.product_name,
             body.product_url, body.product_thumb, body.keywords, int(body.enabled), body.room_id),
        )
        await db.commit()
    return {"id": cur.lastrowid, **body.model_dump()}


@app.post("/api/products/bulk", status_code=201)
async def bulk_create_products(body: list[ProductCreate]):
    ids = []
    async with aio_connect() as db:
        for p in body:
            cur = await db.execute(
                """INSERT INTO products (platform, product_id, product_name, product_url, product_thumb, keywords, enabled, room_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (p.platform, p.product_id, p.product_name,
                 p.product_url, p.product_thumb, p.keywords, int(p.enabled), p.room_id),
            )
            ids.append(cur.lastrowid)
        await db.commit()
    return {"created": len(ids), "ids": ids}


@app.patch("/api/products/{product_id}")
async def update_product(product_id: int, body: ProductUpdate):
    async with aio_connect() as db:
        async with db.execute("SELECT id FROM products WHERE id = ?", (product_id,)) as cur:
            if not await cur.fetchone():
                raise HTTPException(status_code=404, detail="Product not found")
        updates = {k: v for k, v in body.model_dump().items() if v is not None}
        if not updates:
            return {"id": product_id}
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        await db.execute(
            f"UPDATE products SET {set_clause} WHERE id = ?",
            list(updates.values()) + [product_id],
        )
        await db.commit()
    return {"id": product_id, **updates}


@app.delete("/api/products/{product_id}", status_code=204)
async def delete_product(product_id: int):
    async with aio_connect() as db:
        await db.execute("DELETE FROM products WHERE id = ?", (product_id,))
        await db.commit()


@app.get("/api/products/check-url")
async def check_product_url(url: str):
    """Check if a product_url already exists. Returns matched products."""
    if not url or not url.strip():
        return {"duplicates": []}
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT id, product_name, product_url FROM products WHERE product_url = ?",
            (url.strip(),),
        ) as cur:
            rows = await cur.fetchall()
    return {"duplicates": [dict(r) for r in rows]}


@app.get("/api/products/duplicate-urls")
async def list_duplicate_urls():
    """Return all products whose product_url appears more than once."""
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            """SELECT p.id, p.product_name, p.product_url, rm.name as room_name
               FROM products p
               LEFT JOIN rooms rm ON p.room_id = rm.id
               WHERE p.product_url IS NOT NULL AND p.product_url != ''
                 AND p.product_url IN (
                     SELECT product_url FROM products
                     WHERE product_url IS NOT NULL AND product_url != ''
                     GROUP BY product_url HAVING COUNT(*) > 1
                 )
               ORDER BY p.product_url, p.id"""
        ) as cur:
            rows = await cur.fetchall()
    return [dict(r) for r in rows]


@app.post("/api/products/import-excel", status_code=201)
async def import_products_excel(file: UploadFile = File(...)):
    """Import products from an Excel file (.xlsx/.xls).
    Columns: product_name(必填), product_id, product_url, keywords, platform, room_id
    Returns created count and list of skipped rows (missing product_name).
    """
    import openpyxl, io
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法解析Excel文件: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="Excel文件为空")

    # First row = header
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    col = {name: idx for idx, name in enumerate(header)}

    def get(row, key, default=""):
        idx = col.get(key)
        if idx is None:
            return default
        v = row[idx] if idx < len(row) else None
        return str(v).strip() if v is not None else default

    created_ids = []
    skipped = []
    async with aio_connect() as db:
        for i, row in enumerate(rows[1:], start=2):
            name = get(row, "product_name") or get(row, "商品名称")
            if not name:
                skipped.append(i)
                continue
            pid    = get(row, "product_id")    or get(row, "平台商品id") or get(row, "商品id")
            url    = get(row, "product_url")   or get(row, "商品链接")
            thumb  = get(row, "product_thumb") or get(row, "缩略图")
            kw     = get(row, "keywords")      or get(row, "匹配关键词") or get(row, "关键词")
            plat   = get(row, "platform")      or get(row, "平台") or "douyin"
            rid_s  = get(row, "room_id")       or get(row, "直播间id")
            try:
                rid = int(float(rid_s)) if rid_s else None
            except Exception:
                rid = None
            cur2 = await db.execute(
                """INSERT INTO products (platform, product_id, product_name, product_url, product_thumb, keywords, enabled, room_id)
                   VALUES (?, ?, ?, ?, ?, ?, 1, ?)""",
                (plat, pid or None, name, url or None, thumb or None, kw or None, rid),
            )
            created_ids.append(cur2.lastrowid)
        await db.commit()

    return {"created": len(created_ids), "skipped_rows": skipped, "ids": created_ids}


@app.get("/api/products/template.xlsx")
async def download_products_template():
    """Return an Excel template for bulk product import."""
    import openpyxl, io
    from fastapi.responses import Response as FastAPIResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品导入模板"
    headers = ["product_name", "product_id", "product_url", "product_thumb", "keywords", "platform", "room_id"]
    display = ["商品名称(必填)", "平台商品ID", "商品链接", "缩略图URL", "匹配关键词(逗号分隔)", "平台(默认douyin)", "直播间ID"]
    ws.append(headers)
    ws.append(display)
    # Example row
    ws.append(["蓬松波波头假发", "7123456789", "https://haohuo.jinritemai.com/...", "https://p3-aio.ecombdimg.com/img/xxx.jpg", "假发,波波头,黑色", "douyin", "1"])
    # Column widths
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FastAPIResponse(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_template.xlsx"},
    )


# ── Publish Accounts ──────────────────────────────────────────────────────────

COOKIES_DIR = os.path.expanduser("~/.douyin-publisher/cookies")


@app.get("/api/publish-accounts")
async def list_publish_accounts():
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM publish_accounts ORDER BY created_at DESC") as cur:
            rows = await cur.fetchall()
    return [dict(r) for r in rows]


@app.post("/api/publish-accounts", status_code=201)
async def create_publish_account(body: PublishAccountCreate):
    async with aio_connect() as db:
        cur = await db.execute(
            "INSERT INTO publish_accounts (platform, account_name) VALUES (?, ?)",
            (body.platform, body.account_name),
        )
        await db.commit()
        account_id = cur.lastrowid
    return {"id": account_id, "platform": body.platform, "account_name": body.account_name}


@app.delete("/api/publish-accounts/{account_id}", status_code=204)
async def delete_publish_account(account_id: int):
    async with aio_connect() as db:
        await db.execute("DELETE FROM publish_accounts WHERE id = ?", (account_id,))
        await db.commit()


@app.post("/api/publish-accounts/{account_id}/login")
async def login_publish_account(account_id: int):
    """Launch a headed Playwright browser for the user to log in manually."""
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM publish_accounts WHERE id = ?", (account_id,)) as cur:
            account = await cur.fetchone()
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")

    platform = account["platform"]
    os.makedirs(COOKIES_DIR, exist_ok=True)
    cookie_file = os.path.join(COOKIES_DIR, f"{platform}_{account_id}.json")

    try:
        if platform == "douyin":
            from publisher_douyin import DouyinPublisher
            publisher = DouyinPublisher()
        else:
            raise HTTPException(status_code=400, detail=f"Login not supported for platform: {platform}")

        # Run login in background task so it doesn't block the request
        asyncio.create_task(_do_login(publisher, dict(account), cookie_file, account_id))
        return {"account_id": account_id, "cookie_file": cookie_file, "status": "login_started"}
    except ImportError:
        raise HTTPException(status_code=500, detail="playwright not installed")


async def _do_login(publisher, account: dict, cookie_file: str, account_id: int):
    try:
        success = await publisher.login_interactive(account, cookie_file)
    except Exception as e:
        logger.error(f"_do_login unexpected error: {e}")
        await broadcast({"type": "login_done", "account_id": account_id, "success": False})
        return
    if success:
        async with aio_connect() as db:
            await db.execute(
                "UPDATE publish_accounts SET cookie_file = ? WHERE id = ?",
                (cookie_file, account_id),
            )
            await db.commit()
        await broadcast({"type": "login_done", "account_id": account_id, "success": True})
    else:
        await broadcast({"type": "login_done", "account_id": account_id, "success": False})


# ── Publish Tasks ─────────────────────────────────────────────────────────────


@app.get("/api/publish-accounts/{account_id}/check-cookie")
async def check_account_cookie(account_id: int):
    """Check if the account's saved cookies are still valid (not expired/logged-out)."""
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM publish_accounts WHERE id = ?", (account_id,)) as cur:
            account = await cur.fetchone()
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")

    account_dict = dict(account)
    platform = account_dict.get("platform", "douyin")

    if platform == "douyin":
        from publisher_douyin import DouyinPublisher
        publisher = DouyinPublisher()
        try:
            valid = await publisher.login_check(account_dict)
        except Exception as e:
            logger.warning(f"Cookie check failed for account {account_id}: {e}")
            # 检测失败时假定有效，不阻止排期流程
            return {"account_id": account_id, "valid": True, "error": str(e)}
        return {"account_id": account_id, "valid": valid}
    else:
        return {"account_id": account_id, "valid": False, "error": f"Unsupported platform: {platform}"}


@app.get("/api/publish-tasks")
async def list_publish_tasks(status: Optional[str] = None):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        if status:
            async with db.execute(
                """SELECT t.*, g.label as group_label, g.merged_filename,
                          pa.account_name, pa.platform as account_platform,
                          p.product_name, t.product_ids, rm.name as room_name,
                          t.manual_published, t.manual_published_at
                   FROM publish_tasks t
                   JOIN clip_groups g ON t.group_id = g.id
                   LEFT JOIN rooms rm ON g.room_id = rm.id
                   LEFT JOIN publish_accounts pa ON t.account_id = pa.id
                   LEFT JOIN products p ON t.product_id = p.id
                   WHERE t.status = ?
                   ORDER BY t.created_at DESC""",
                (status,),
            ) as cur:
                rows = await cur.fetchall()
        else:
            async with db.execute(
                """SELECT t.*, g.label as group_label, g.merged_filename,
                          pa.account_name, pa.platform as account_platform,
                          p.product_name, t.product_ids, rm.name as room_name,
                          t.manual_published, t.manual_published_at
                   FROM publish_tasks t
                   JOIN clip_groups g ON t.group_id = g.id
                   LEFT JOIN rooms rm ON g.room_id = rm.id
                   LEFT JOIN publish_accounts pa ON t.account_id = pa.id
                   LEFT JOIN products p ON t.product_id = p.id
                   ORDER BY t.created_at DESC"""
            ) as cur:
                rows = await cur.fetchall()
    return [dict(r) for r in rows]


@app.post("/api/publish-tasks", status_code=201)
async def create_publish_task(body: PublishTaskCreate):
    # Validate group exists and has a merged video
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM clip_groups WHERE id = ?", (body.group_id,)) as cur:
            group = await cur.fetchone()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    from video_path_resolver import resolve_publish_video
    requested_version = group["publish_versions"] or "both"
    video_path, selected_version, checked_paths, available_versions = resolve_publish_video(
        dict(group), requested_version
    )
    if not video_path:
        checked = "; ".join(checked_paths) or "无候选路径"
        alternatives = ", ".join(available_versions) or "无"
        raise HTTPException(
            status_code=409,
            detail=(
                f"group_id={body.group_id} 选择版本={requested_version} 无可用视频文件；"
                f"已检查候选：{checked}；可用替代版本：{alternatives}"
            ),
        )

    # Duplicate publish guard: same group + platform already has an active or completed task
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            """SELECT id, status FROM publish_tasks
               WHERE group_id=? AND platform=? AND status IN ('pending','scheduled','publishing','done')
               LIMIT 1""",
            (body.group_id, body.platform),
        ) as cur:
            existing = await cur.fetchone()
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"duplicate: group {body.group_id} already has a {existing['status']} task for {body.platform} (task_id={existing['id']})"
        )

    title = body.title
    description = body.description
    tags = body.tags

    # Auto-generate metadata via LLM if requested, OR if title is empty
    if body.auto_meta or (not title):
        meta = await generate_meta(body.group_id)
        if meta:
            title = title or meta.get("title")
            description = description or meta.get("description")
            tags = tags or meta.get("tags")
        # Fallback if still no title
        if not title:
            async with aio_connect() as db:
                db.row_factory = aiosqlite.Row
                async with db.execute("SELECT label FROM clip_groups WHERE id = ?", (body.group_id,)) as cur:
                    grp = await cur.fetchone()
                title = grp["label"] if grp else f"分组 {body.group_id}"
                description = ""
                tags = ""

    status = "scheduled" if body.scheduled_at else "pending"
    product_ids_str = ",".join(str(i) for i in body.product_ids) if body.product_ids else None
    # keep product_id as first item for backward compat
    first_product_id = body.product_ids[0] if body.product_ids else body.product_id

    async with aio_connect() as db:
        cur = await db.execute(
            """INSERT INTO publish_tasks
               (group_id, platform, account_id, status, scheduled_at,
                title, description, tags, product_id, product_ids, video_path, no_cart)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (body.group_id, body.platform, body.account_id, status,
             body.scheduled_at, title, description, tags,
             first_product_id, product_ids_str, video_path, 1 if body.no_cart else 0),
        )
        await db.commit()
        task_id = cur.lastrowid

    return {
        "id": task_id,
        "group_id": body.group_id,
        "platform": body.platform,
        "status": status,
        "title": title,
        "description": description,
        "tags": tags,
    }


@app.get("/api/publish-tasks/unscheduled-groups")
async def get_unscheduled_groups(platform: str = "douyin", room_id: Optional[int] = None):
    """
    Return merged groups that have no active/done publish task for the given platform.
    Used by the batch-schedule UI to preview how many groups will be queued.
    """
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        sql = """
            SELECT g.id, g.label, g.merged_filename, g.director_final_video,
                   g.realistic_final_video, g.conservative_final_video,
                   g.creative_final_video, g.qianchuan_final_video, g.publish_versions, g.room_id, rm.name as room_name
            FROM clip_groups g
            LEFT JOIN rooms rm ON g.room_id = rm.id
            WHERE (g.merge_status = 2 OR g.classic_status = 2 OR g.director_status = 2 OR g.realistic_status = 2 OR g.conservative_status = 2 OR g.creative_status = 2 OR g.qianchuan_status = 2)
              AND (g.merged_filename IS NOT NULL OR g.director_final_video IS NOT NULL OR g.realistic_final_video IS NOT NULL OR g.conservative_final_video IS NOT NULL OR g.creative_final_video IS NOT NULL OR g.qianchuan_final_video IS NOT NULL)
              AND g.label != '未分类'
              AND NOT EXISTS (
                  SELECT 1 FROM publish_tasks pt
                  WHERE pt.group_id = g.id
                    AND pt.platform = ?
                    AND pt.status IN ('pending','scheduled','publishing','done')
              )
        """
        params: list = [platform]
        if room_id:
            sql += " AND g.room_id = ?"
            params.append(room_id)
        async with db.execute(sql, params) as cur:
            rows = await cur.fetchall()
    available_rows = []
    for row in rows:
        item = dict(row)
        item.update(_artifact_statuses(item))
        item.update(_publish_version_metadata(item))
        item["available_versions"] = [version for version in PUBLISHABLE_VERSIONS if item.get(f"{version}_available")]
        item["video_available"] = bool(item["available_versions"])
        item["missing_reason"] = None if item["video_available"] else "not_generated"
        if item["video_available"]:
            available_rows.append(item)
    return available_rows


@app.post("/api/publish-tasks/batch-schedule", status_code=201)
async def batch_schedule_tasks(body: BatchScheduleCreate):
    """
    Create scheduled publish tasks for all unscheduled merged groups.
    Tasks are spaced by interval_minutes starting from start_datetime.
    """
    from datetime import datetime, timedelta

    # Parse start time
    try:
        start_dt = datetime.fromisoformat(body.start_datetime)
        end_dt = datetime.fromisoformat(body.end_datetime) if body.end_datetime else None
    except ValueError:
        raise HTTPException(status_code=422, detail="start_datetime and end_datetime must be ISO format, e.g. 2026-03-25T10:00:00")
    if end_dt and end_dt < start_dt:
        raise HTTPException(status_code=422, detail="end_datetime must not be earlier than start_datetime")

    # Fetch eligible groups (same logic as unscheduled-groups endpoint)
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        sql = """
            SELECT g.id, g.label, g.merged_filename, g.director_final_video,
                   g.creative_final_video, g.qianchuan_final_video, g.publish_versions, g.room_id
            FROM clip_groups g
            WHERE (g.merge_status = 2 OR g.classic_status = 2 OR g.director_status = 2 OR g.realistic_status = 2 OR g.conservative_status = 2 OR g.creative_status = 2 OR g.qianchuan_status = 2)
              AND (g.merged_filename IS NOT NULL OR g.director_final_video IS NOT NULL OR g.realistic_final_video IS NOT NULL OR g.conservative_final_video IS NOT NULL OR g.creative_final_video IS NOT NULL OR g.qianchuan_final_video IS NOT NULL)
              AND g.label != '未分类'
              AND NOT EXISTS (
                  SELECT 1 FROM publish_tasks pt
                  WHERE pt.group_id = g.id
                    AND pt.platform = ?
                    AND pt.status IN ('pending','scheduled','publishing','done')
              )
        """
        params: list = [body.platform]
        if body.room_id:
            sql += " AND g.room_id = ?"
            params.append(body.room_id)
        sql += " ORDER BY g.created_at ASC"
        async with db.execute(sql, params) as cur:
            groups = await cur.fetchall()

    # 过滤排除的分组，并尊重前端按时间槽位自动选择的分组。
    if body.include_group_ids is not None:
        include_set = set(body.include_group_ids)
        groups = [g for g in groups if g["id"] in include_set]
    if body.exclude_group_ids:
        exclude_set = set(body.exclude_group_ids)
        groups = [g for g in groups if g["id"] not in exclude_set]

    if not groups:
        return {"created": 0, "tasks": [], "message": "没有找到可排期的分组"}

    product_ids_str = ",".join(str(i) for i in body.product_ids) if body.product_ids else None
    first_product_id = body.product_ids[0] if body.product_ids else None

    # Keep backend selection count aligned with the UI's usable publishing window.
    # A batch beginning before 09:00 reserves the first four hours for preparation;
    # the requested example (05:00-21:00 every 20 minutes) therefore has 36 slots.
    slot_limit = len(groups)
    if end_dt:
        total_minutes = int((end_dt - start_dt).total_seconds() // 60)
        preparation_minutes = 4 * 60 if start_dt.hour < 9 <= end_dt.hour else 0
        slot_limit = max(0, (total_minutes - preparation_minutes) // body.interval_minutes)
    groups_to_process = groups[:slot_limit]

    video_base = os.path.join(os.path.dirname(__file__), "..", "recordings")
    created_tasks = []

    # ── 黄金时段对齐辅助函数 ───────────────────────────────────────────
    def _snap_to_golden_hour(dt: datetime) -> datetime:
        """将时间对齐到最近的黄金时段起始点。
        黄金时段：12:00-13:00 / 18:00-20:00 / 21:00-22:00
        如果已在黄金时段内，不调整。
        如果在低活跃时段（00:00-07:00），顺延到当天 12:00。
        """
        # 黄金时段区间 (start_hour, end_hour)
        GOLDEN = [(12, 13), (18, 20), (21, 22)]
        h = dt.hour
        # 已在黄金时段内
        for gs, ge in GOLDEN:
            if gs <= h < ge:
                return dt
        # 低活跃时段 00-07，跳到当天 12:00
        if 0 <= h < 7:
            return dt.replace(hour=12, minute=0, second=0, microsecond=0)
        # 其他时段，找下一个黄金时段起始点
        from datetime import date as _date
        for gs, ge in GOLDEN:
            if h < gs:
                return dt.replace(hour=gs, minute=dt.minute % 60, second=0, microsecond=0)
        # 超过最晚黄金时段，顺延到明天 12:00
        next_day = dt.replace(hour=12, minute=0, second=0, microsecond=0) + timedelta(days=1)
        return next_day
    # ────────────────────────────────────────────────────────────────────

    # Phase 1: insert tasks immediately without waiting for LLM meta
    async with aio_connect() as db:
        for i, group in enumerate(groups_to_process):
            raw_dt = start_dt + timedelta(minutes=body.interval_minutes * i)
            scheduled_dt = _snap_to_golden_hour(raw_dt)
            # Do not create a task after the user-selected end time. Apply the
            # limit after golden-hour snapping, since that is the actual publish time.
            if end_dt and scheduled_dt > end_dt:
                break
            scheduled_at = scheduled_dt.isoformat()
            from video_path_resolver import resolve_publish_video
            video_path, selected_version, checked_paths, available_versions = resolve_publish_video(
                dict(group), group["publish_versions"] or "both"
            )
            if not video_path:
                logger.warning(
                    "Skipping batch group %s: no available video; checked=%s alternatives=%s",
                    group["id"], checked_paths, available_versions,
                )
                continue

            cur = await db.execute(
                """INSERT INTO publish_tasks
                   (group_id, platform, account_id, status, scheduled_at,
                    title, description, tags, product_id, product_ids, video_path, no_cart)
                   VALUES (?, ?, ?, 'scheduled', ?, ?, ?, ?, ?, ?, ?, ?)""",
                (group["id"], body.platform, body.account_id, scheduled_at,
                 None, None, None,
                 first_product_id, product_ids_str, video_path,
                 1 if body.no_cart else 0),
            )
            created_tasks.append({
                "task_id": cur.lastrowid,
                "group_id": group["id"],
                "group_label": group["label"],
                "scheduled_at": scheduled_at,
            })
        await db.commit()

    # Phase 2: if auto_meta requested, generate titles in background (non-blocking)
    if body.auto_meta and created_tasks:
        async def _fill_meta_background(task_list):
            for item in task_list:
                try:
                    meta = await generate_meta(item["group_id"])
                    if not meta:
                        logger.warning(f"[meta] generate_meta returned None for group {item['group_id']} (task {item['task_id']})")
                        continue
                    # generate_meta returns {"schemes": [...]} — pick first scheme
                    schemes = meta.get("schemes", [])
                    if schemes:
                        best = schemes[0]
                    else:
                        best = meta  # legacy single-scheme
                    title = best.get("title") or meta.get("title")
                    description = best.get("description") or meta.get("description")
                    tags = best.get("tags") or meta.get("tags")
                    if not title:
                        logger.warning(f"[meta] No title generated for group {item['group_id']} (task {item['task_id']}), using fallback")
                        # Fallback: use group label as title
                        title = item.get("group_label", f"分组 {item['group_id']}")
                        description = ""
                        tags = ""
                    async with aio_connect() as db2:
                        await db2.execute(
                            "UPDATE publish_tasks SET title=?, description=?, tags=? WHERE id=?",
                            (title, description, tags, item["task_id"]),
                        )
                        await db2.commit()
                except Exception as e:
                    logger.error(f"[meta] Failed to fill meta for task {item.get('task_id')}: {e}")
        asyncio.create_task(_fill_meta_background(list(created_tasks)))

    return {
        "created": len(created_tasks),
        "tasks": created_tasks,
        "message": f"已为 {len(created_tasks)} 个分组创建排期任务",
    }


@app.get("/api/publish-tasks/{task_id}")
async def get_publish_task(task_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            """SELECT t.*, g.label as group_label, g.merged_filename,
                      pa.account_name, p.product_name
               FROM publish_tasks t
               JOIN clip_groups g ON t.group_id = g.id
               LEFT JOIN publish_accounts pa ON t.account_id = pa.id
               LEFT JOIN products p ON t.product_id = p.id
               WHERE t.id = ?""",
            (task_id,),
        ) as cur:
            task = await cur.fetchone()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return dict(task)


@app.delete("/api/publish-tasks/bulk-cancel", status_code=200)
async def bulk_cancel_publish_tasks(body: dict):
    """Bulk cancel pending/scheduled tasks. body: {status: 'all'|'scheduled'|'pending', ids: [int, ...]}
    If ids is provided, cancel only those. Otherwise cancel by status filter."""
    ids = body.get("ids")
    status_filter = body.get("status", "all")
    async with aio_connect() as db:
        if ids:
            placeholders = ",".join("?" * len(ids))
            async with db.execute(
                f"SELECT id FROM publish_tasks WHERE id IN ({placeholders}) AND status IN ('pending','scheduled','failed')",
                ids,
            ) as cur:
                valid_ids = [r[0] for r in await cur.fetchall()]
            if valid_ids:
                await db.execute(
                    f"DELETE FROM publish_tasks WHERE id IN ({','.join('?' * len(valid_ids))})",
                    valid_ids,
                )
        else:
            if status_filter == "all":
                await db.execute("DELETE FROM publish_tasks WHERE status IN ('pending','scheduled','failed')")
            elif status_filter in ("pending", "scheduled", "failed"):
                await db.execute("DELETE FROM publish_tasks WHERE status = ?", (status_filter,))
            else:
                raise HTTPException(status_code=400, detail="status must be all/scheduled/pending/failed")
        await db.commit()
        async with db.execute("SELECT changes()") as cur:
            deleted = (await cur.fetchone())[0]
    return {"deleted": deleted}


@app.delete("/api/publish-tasks/{task_id}", status_code=204)
async def cancel_publish_task(task_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT status FROM publish_tasks WHERE id = ?", (task_id,)) as cur:
            task = await cur.fetchone()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    if task["status"] not in ("pending", "scheduled", "failed"):
        raise HTTPException(status_code=409, detail="Can only cancel pending/scheduled/failed tasks")
    async with aio_connect() as db:
        await db.execute("DELETE FROM publish_tasks WHERE id = ?", (task_id,))
        await db.commit()


@app.post("/api/publish-tasks/{task_id}/retry")
async def retry_publish_task(task_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM publish_tasks WHERE id = ?", (task_id,)) as cur:
            task = await cur.fetchone()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    if task["status"] not in ("failed", "scheduled"):
        raise HTTPException(status_code=409, detail="Can only retry failed or scheduled tasks")
    async with aio_connect() as db:
        await db.execute(
            "UPDATE publish_tasks SET status = 'pending', error_msg = NULL WHERE id = ?",
            (task_id,),
        )
        await db.commit()
    return {"task_id": task_id, "status": "pending"}


@app.patch("/api/publish-tasks/{task_id}")
async def update_publish_task(task_id: int, body: dict):
    """Partial update: currently supports scheduled_at (reschedule) and status reset."""
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM publish_tasks WHERE id = ?", (task_id,)) as cur:
            task = await cur.fetchone()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    updates = {}
    if "scheduled_at" in body:
        updates["scheduled_at"] = body["scheduled_at"]
        # Reset status to 'scheduled' when rescheduling an expired task
        if task["status"] in ("scheduled", "failed"):
            updates["status"] = "scheduled"
            updates["error_msg"] = None
    if not updates:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    async with aio_connect() as db:
        await db.execute(
            f"UPDATE publish_tasks SET {set_clause} WHERE id = ?",
            (*updates.values(), task_id),
        )
        await db.commit()
    return {"task_id": task_id, **updates}


@app.post("/api/publish-tasks/{task_id}/regen-meta")
async def regen_publish_task_meta(task_id: int):
    """Re-generate title/description/tags for a single publish task via LLM."""
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT group_id FROM publish_tasks WHERE id = ?", (task_id,)) as cur:
            task = await cur.fetchone()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    meta = await generate_meta(task["group_id"])
    if not meta:
        raise HTTPException(status_code=500, detail="LLM metadata generation failed")
    parsed = _extract_publish_meta(meta)
    title = parsed["title"]
    description = parsed["description"]
    tags = parsed["tags"]
    if not title:
        raise HTTPException(status_code=500, detail="No title generated")
    async with aio_connect() as db:
        await db.execute(
            "UPDATE publish_tasks SET title=?, description=?, tags=? WHERE id=?",
            (title, description, tags, task_id),
        )
        await db.commit()
    return {"task_id": task_id, "title": title, "description": description, "tags": tags}


def _extract_publish_meta(meta: Optional[dict], fallback_title: str = "") -> dict:
    """Normalize meta_generator output to title/description/tags with safe fallbacks."""
    meta = meta or {}
    schemes = meta.get("schemes", []) if isinstance(meta, dict) else []
    best = schemes[0] if schemes else meta
    title = (best.get("title") or meta.get("title") or fallback_title or "").strip()
    description = (best.get("description") or meta.get("description") or "").strip()
    tags = (best.get("tags") or meta.get("tags") or "").strip()
    return {"title": title, "description": description, "tags": tags}


@app.post("/api/publish-tasks/bulk-regen-meta")
async def bulk_regen_publish_task_meta(body: dict):
    """Batch-generate title/description/tags for pending/scheduled publish tasks.

    Default behavior fills only tasks missing title or description, so existing manually-edited
    publish copy is not overwritten. Pass {force: true} to regenerate selected tasks anyway.
    """
    task_ids = body.get("task_ids")  # None = all pending/scheduled tasks missing metadata
    force = bool(body.get("force", False))
    statuses = body.get("statuses") or ["pending", "scheduled"]
    statuses = [s for s in statuses if s in {"pending", "scheduled", "failed"}]
    if not statuses:
        statuses = ["pending", "scheduled"]

    where = [f"status IN ({','.join('?' * len(statuses))})"]
    params: list = list(statuses)

    if task_ids:
        task_ids = [int(i) for i in task_ids]
        where.append(f"id IN ({','.join('?' * len(task_ids))})")
        params.extend(task_ids)
    if not force:
        where.append("(title IS NULL OR TRIM(title) = '' OR description IS NULL OR TRIM(description) = '')")

    where_sql = " AND ".join(where)
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            f"""SELECT t.id, t.group_id, g.label as group_label
                FROM publish_tasks t
                JOIN clip_groups g ON t.group_id = g.id
                WHERE {where_sql}
                ORDER BY t.scheduled_at ASC, t.created_at ASC""",
            params,
        ) as cur:
            tasks = await cur.fetchall()

    if not tasks:
        return {"updated": 0, "skipped": 0, "total": 0, "message": "没有需要生成文案的待发布/定时任务"}

    updated = 0
    skipped = 0
    updated_ids = []
    for task in tasks:
        try:
            meta = await generate_meta(task["group_id"])
            parsed = _extract_publish_meta(meta, task["group_label"] or f"分组 {task['group_id']}")
            title = parsed["title"]
            description = parsed["description"]
            tags = parsed["tags"]
            if not title:
                skipped += 1
                continue
            async with aio_connect() as db2:
                await db2.execute(
                    "UPDATE publish_tasks SET title=?, description=?, tags=? WHERE id=?",
                    (title, description, tags, task["id"]),
                )
                await db2.commit()
            updated += 1
            updated_ids.append(task["id"])
        except Exception as e:
            logger.error(f"[bulk-meta] Failed for task {task['id']}: {e}")
            skipped += 1

    return {"updated": updated, "skipped": skipped, "total": len(tasks), "task_ids": updated_ids}


# ── Manual publish mark ───────────────────────────────────────────────────────

@app.post("/api/publish-tasks/{task_id}/manual-publish")
async def mark_manual_publish(task_id: int):
    """Mark a publish task as manually published (user downloaded and published externally)."""
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT id, status FROM publish_tasks WHERE id = ?", (task_id,)) as cur:
            task = await cur.fetchone()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    now = datetime.now().isoformat()
    async with aio_connect() as db:
        await db.execute(
            "UPDATE publish_tasks SET status = 'done', manual_published = 1, manual_published_at = ? WHERE id = ?",
            (now, task_id),
        )
        await db.commit()
    return {"task_id": task_id, "status": "done", "manual_published": True, "manual_published_at": now}


# ── Meta generation ───────────────────────────────────────────────────────────

@app.post("/api/groups/{group_id}/generate-meta")
async def generate_group_meta(group_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT id FROM clip_groups WHERE id = ?", (group_id,)) as cur:
            if not await cur.fetchone():
                raise HTTPException(status_code=404, detail="Group not found")
    meta = await generate_meta(group_id)
    if not meta:
        raise HTTPException(status_code=500, detail="LLM metadata generation failed")
    return meta


# ── Group thumbnail regeneration ──────────────────────────────────────────────

@app.post("/api/groups/{group_id}/generate-thumbnail")
async def generate_group_thumbnail(group_id: int, body: dict = {}):
    """Regenerate the thumbnail for a group's merged video with a specific scheme."""
    scheme_type = body.get("scheme_type", "种草") if body else "种草"
    title = body.get("title", "") if body else ""
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT merged_filename FROM clip_groups WHERE id = ?", (group_id,)
        ) as cur:
            group = await cur.fetchone()
    if not group or not group["merged_filename"]:
        raise HTTPException(status_code=404, detail="Group or merged video not found")
    mp4_path = os.path.join(RECORDINGS_DIR, group["merged_filename"])
    if not os.path.exists(mp4_path):
        raise HTTPException(status_code=404, detail="Merged video file not found")
    thumb = await generate_thumbnail(mp4_path, title=title or "假发变美瞬间", scheme_type=scheme_type)
    if not thumb:
        raise HTTPException(status_code=500, detail="Thumbnail generation failed")
    thumb_rel = os.path.relpath(thumb, RECORDINGS_DIR)
    return {"thumbnail": thumb_rel, "scheme_type": scheme_type}


@app.post("/api/groups/{group_id}/generate-covers")
async def generate_group_covers(group_id: int):
    """Generate 6 cover candidates using different marketing schemes.
    Frames are extracted from the longest *original* recording (no burned-in subtitles).
    SRT text is passed to enable content-aware dynamic titles.
    """
    from thumbnail import generate_cover_candidates
    import json as _json
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT merged_filename FROM clip_groups WHERE id = ?",
            (group_id,),
        ) as cur:
            group = await cur.fetchone()
        if not group:
            raise HTTPException(status_code=404, detail="Group not found")
        # Use the longest original recording (no subtitles burned in)
        async with db.execute(
            """SELECT filename FROM recordings
               WHERE group_id = ? AND filename IS NOT NULL
               ORDER BY (end_time - start_time) DESC LIMIT 1""",
            (group_id,),
        ) as cur:
            orig_row = await cur.fetchone()

    # Prefer original recording (subtitle-free) → fall back to classic merged video
    mp4_path = None
    if orig_row and orig_row["filename"]:
        candidate = os.path.join(RECORDINGS_DIR, orig_row["filename"])
        if os.path.exists(candidate):
            mp4_path = candidate
    if not mp4_path and group["merged_filename"]:
        mp4_path = os.path.join(RECORDINGS_DIR, group["merged_filename"])
    if not mp4_path or not os.path.exists(mp4_path):
        raise HTTPException(status_code=404, detail="No video available for cover generation")

    # Load merged SRT for dynamic title generation
    srt_text = ""
    if group["merged_filename"]:
        srt_path = os.path.join(RECORDINGS_DIR, os.path.splitext(group["merged_filename"])[0] + ".srt")
        if os.path.exists(srt_path):
            try:
                with open(srt_path, "r", encoding="utf-8", errors="ignore") as _f:
                    srt_text = _f.read()
            except Exception:
                pass

    covers_dir = os.path.join(RECORDINGS_DIR, "covers")
    candidates = await generate_cover_candidates(mp4_path, group_id, covers_dir, srt_text=srt_text)
    if not candidates:
        raise HTTPException(status_code=500, detail="Cover generation failed")

    rel_paths = [os.path.relpath(p, RECORDINGS_DIR) for p in candidates]
    async with aio_connect() as db:
        await db.execute(
            "UPDATE clip_groups SET cover_candidates = ? WHERE id = ?",
            (_json.dumps(rel_paths, ensure_ascii=False), group_id),
        )
        await db.commit()
    return {"group_id": group_id, "covers": rel_paths}


@app.post("/api/groups/{group_id}/select-cover")
async def select_group_cover(group_id: int, body: dict):
    """Set the selected cover for a group."""
    cover = body.get("cover")
    if not cover:
        raise HTTPException(status_code=400, detail="cover field required")
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT cover_candidates FROM clip_groups WHERE id = ?", (group_id,)
        ) as cur:
            group = await cur.fetchone()
        if not group:
            raise HTTPException(status_code=404, detail="Group not found")
        try:
            candidates = json.loads(group["cover_candidates"] or "[]")
        except (TypeError, json.JSONDecodeError) as exc:
            raise HTTPException(status_code=409, detail="Cover candidates are invalid; regenerate covers") from exc
        if not isinstance(candidates, list) or cover not in candidates:
            raise HTTPException(status_code=400, detail="cover must be one of the generated candidates")
        await db.execute(
            "UPDATE clip_groups SET selected_cover = ? WHERE id = ?",
            (cover, group_id),
        )
        await db.commit()
    return {"group_id": group_id, "selected_cover": cover}


@app.get("/api/groups/{group_id}/cover/{filename:path}")
async def get_group_cover(group_id: int, filename: str):
    """Serve a cover candidate image."""
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT cover_candidates FROM clip_groups WHERE id = ?", (group_id,)
        ) as cur:
            group = await cur.fetchone()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    try:
        candidates = json.loads(group["cover_candidates"] or "[]")
    except (TypeError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=409, detail="Cover candidates are invalid; regenerate covers") from exc
    if not isinstance(candidates, list) or filename not in candidates:
        raise HTTPException(status_code=404, detail="Cover not associated with group")
    cover_root = os.path.realpath(RECORDINGS_DIR)
    cover_path = os.path.realpath(os.path.join(cover_root, filename))
    if not cover_path.startswith(cover_root + os.sep) or not os.path.isfile(cover_path):
        raise HTTPException(status_code=404, detail="Cover not found")
    return FileResponse(
        cover_path,
        media_type="image/jpeg",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )


# ── Product matching ──────────────────────────────────────────────────────────

@app.post("/api/groups/{group_id}/match-product")
async def match_group_product(group_id: int):
    async with aio_connect() as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT id FROM clip_groups WHERE id = ?", (group_id,)) as cur:
            if not await cur.fetchone():
                raise HTTPException(status_code=404, detail="Group not found")
    product = await match_product(group_id)
    return {"group_id": group_id, "product": product}




# ── Stream login status ───────────────────────────────────────────────────────

import glob as _glob
import time as _time

_STREAM_COOKIE_DIR = os.path.expanduser("~/.douyin-publisher/cookies")
_STREAM_AUTH_KEYS  = {"sessionid", "uid_tt", "sid_guard"}

# Track ongoing refresh so we don't launch two browsers
_stream_login_task: Optional[asyncio.Task] = None
_stream_login_state = {"status": "idle", "msg": None, "detail": None}


def _stream_login_diagnostics() -> dict:
    """Best-effort process/session diagnostics, especially useful on Windows RDP."""
    detail = {"pid": os.getpid(), "platform": os.name}
    try:
        import getpass
        detail["process_user"] = getpass.getuser()
    except Exception:
        detail["process_user"] = None
    detail["session_name"] = os.environ.get("SESSIONNAME") or os.environ.get("XDG_SESSION_ID")
    detail["session_id"] = None
    if os.name == "nt":
        try:
            import ctypes
            session_id = ctypes.c_ulong()
            if ctypes.windll.kernel32.ProcessIdToSessionId(os.getpid(), ctypes.byref(session_id)):
                detail["session_id"] = session_id.value
        except Exception as exc:
            detail["session_id_error"] = str(exc)
    detail["interactive_sessions"] = []
    if os.name == "nt":
        try:
            result = __import__("subprocess").run(
                ["quser"], capture_output=True, text=True, timeout=3, creationflags=0x08000000
            )
            # Keep the raw lines as well as parsed fields: quser formatting varies
            # slightly between Windows versions and RDP/console sessions.
            detail["interactive_sessions_raw"] = [line.strip() for line in result.stdout.splitlines() if line.strip()]
            detail["interactive_sessions"] = _parse_quser_sessions(
                detail["interactive_sessions_raw"]
            )
            if result.returncode:
                detail["interactive_sessions_error"] = result.stderr.strip() or f"quser exited {result.returncode}"
        except Exception as exc:
            detail["interactive_sessions_error"] = str(exc)
    return detail


def _parse_quser_sessions(lines: list[str]) -> list[dict]:
    """Parse the stable columns from ``quser`` without relying on padding.

    ``quser`` prefixes the current user with ``>`` and can omit SESSIONNAME
    for disconnected sessions.  Keeping this parser separate makes the
    session decision testable and prevents a formatting variation from
    incorrectly treating session 0 as interactive.
    """
    sessions = []
    for line in lines[1:]:
        fields = line.lstrip("> ").split()
        numeric_index = next(
            (index for index, field in enumerate(fields[:4]) if field.isdigit()),
            None,
        )
        if numeric_index is None or numeric_index == 0 or len(fields) <= numeric_index + 1:
            continue
        session_id = int(fields[numeric_index])
        sessions.append({
            "user": fields[0],
            "session_name": fields[1] if numeric_index >= 2 else None,
            "session_id": session_id,
            "state": fields[numeric_index + 1],
        })
    return sessions


def _stream_login_active_sessions(diagnostics: dict) -> list[dict]:
    """Return sessions where Windows can display a headed browser window."""
    return [
        session for session in diagnostics.get("interactive_sessions", [])
        if str(session.get("state", "")).casefold() in {"active", "运行中"}
    ]


def _stream_login_interactive_session_guard() -> Optional[dict]:
    """Reject headed browser launches from services/non-visible Windows sessions."""
    if os.name != "nt":
        return None
    diagnostics = _stream_login_diagnostics()
    process_session_id = diagnostics.get("session_id")
    process_user = (diagnostics.get("process_user") or "").casefold()
    active_sessions = _stream_login_active_sessions(diagnostics)
    matching = [
        session for session in active_sessions
        if session.get("session_id") == process_session_id
        and str(session.get("user", "")).casefold() == process_user
    ]
    if process_session_id is not None and matching:
        return None
    return {
        "msg": "无法打开登录浏览器：后端不在当前可见的 Windows 交互会话中",
        "detail": "请运行 backend\\start_windows_backend.ps1，或在当前 RDP/桌面会话中启动后端；该启动器会在重新连接时迁移后端",
        "active_sessions": active_sessions,
        "diagnostics": diagnostics,
    }


@app.get("/api/stream-login/status")
async def stream_login_status():
    """Return auth-cookie status and launch diagnostics used for high-quality stream recording."""
    files = sorted(_glob.glob(os.path.join(_STREAM_COOKIE_DIR, "douyin_*.json")))
    refreshing = _stream_login_task is not None and not _stream_login_task.done()
    result = {
        "logged_in": False, "quality": "LD1", "file_age_hours": None,
        "refreshing": refreshing, "diagnostics": _stream_login_diagnostics(),
        "launch_status": _stream_login_state["status"],
    }
    if _stream_login_state["msg"]:
        result["msg"] = _stream_login_state["msg"]
    if _stream_login_state["detail"]:
        result["detail"] = _stream_login_state["detail"]
    if not files:
        result["reason"] = "未找到 Cookie 文件"
        return result

    cookie_file = files[0]
    result["cookie_file"] = os.path.basename(cookie_file)
    result["file_age_hours"] = round((_time.time() - os.path.getmtime(cookie_file)) / 3600, 1)
    try:
        with open(cookie_file, encoding="utf-8") as f:
            cookies = json.load(f)
        names = {c["name"] for c in cookies if "name" in c}
        result["logged_in"] = bool(names & _STREAM_AUTH_KEYS)
    except Exception as exc:
        result["detail"] = f"Cookie 文件读取失败: {exc}"
    result["quality"] = "ORIGIN" if result["logged_in"] else "LD1"
    return result


@app.post("/api/stream-login/refresh")
async def stream_login_refresh():
    """Launch a headed Playwright browser and only report success after it starts."""
    global _stream_login_task, _stream_login_state
    if _stream_login_task and not _stream_login_task.done():
        return {"ok": False, "msg": "登录浏览器已打开，请在浏览器中完成登录", "detail": _stream_login_diagnostics()}
    session_error = _stream_login_interactive_session_guard()
    if session_error:
        _stream_login_state = {"status": "failed", "msg": session_error["msg"], "detail": session_error["detail"]}
        raise HTTPException(status_code=409, detail=session_error)
    try:
        from playwright.async_api import async_playwright
    except ImportError as exc:
        _stream_login_state = {"status": "failed", "msg": "后端未安装 Playwright，无法打开登录浏览器", "detail": "请在运行 8899 后端的 Python 环境执行: pip install playwright && playwright install chromium"}
        logger.error("[stream-login] playwright not installed")
        raise HTTPException(status_code=503, detail={"msg": _stream_login_state["msg"], "detail": _stream_login_state["detail"], "diagnostics": _stream_login_diagnostics()}) from exc

    ready = asyncio.get_running_loop().create_future()
    _stream_login_state = {"status": "starting", "msg": None, "detail": None}

    async def _do_browser_login():
        global _stream_login_state
        browser = None
        try:
            os.makedirs(_STREAM_COOKIE_DIR, exist_ok=True)
            cookie_file = os.path.join(_STREAM_COOKIE_DIR, "douyin_stream.json")
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=False)
                ctx = await browser.new_context()
                page = await ctx.new_page()
                await page.goto("https://live.douyin.com/", timeout=30000, wait_until="domcontentloaded")
                _stream_login_state = {"status": "running", "msg": "登录浏览器已启动，请在可见窗口中完成登录", "detail": _stream_login_diagnostics()}
                if not ready.done():
                    ready.set_result(True)
                logger.info("[stream-login] Browser started; waiting for user to log in (max 5 min)…")
                deadline = _time.time() + 300
                while _time.time() < deadline:
                    cookies = await ctx.cookies("https://live.douyin.com")
                    if {c["name"] for c in cookies if "name" in c} & _STREAM_AUTH_KEYS:
                        with open(cookie_file, "w", encoding="utf-8") as f:
                            json.dump(cookies, f, ensure_ascii=False, indent=2)
                        import douyin_live as _dl
                        _dl._auth_cookies_loaded = False
                        _dl._auth_cookies = {}
                        _stream_login_state = {"status": "completed", "msg": "登录成功，Cookie 已保存", "detail": _stream_login_diagnostics()}
                        break
                    await asyncio.sleep(2)
                else:
                    _stream_login_state = {"status": "timed_out", "msg": "登录超时，未检测到抖音登录 Cookie", "detail": _stream_login_diagnostics()}
                await page.close()
        except Exception as exc:
            _stream_login_state = {"status": "failed", "msg": "登录浏览器启动失败", "detail": str(exc), "diagnostics": _stream_login_diagnostics()}
            logger.exception("[stream-login] Browser error")
            if not ready.done():
                ready.set_exception(exc)
        finally:
            if browser is not None:
                await browser.close()

    _stream_login_task = asyncio.create_task(_do_browser_login())
    try:
        await asyncio.wait_for(asyncio.shield(ready), timeout=35)
    except Exception as exc:
        raise HTTPException(status_code=500, detail={"msg": _stream_login_state["msg"] or "登录浏览器启动失败", "detail": _stream_login_state["detail"] or str(exc), "diagnostics": _stream_login_diagnostics()}) from exc
    return {"ok": True, "msg": _stream_login_state["msg"], "detail": _stream_login_state["detail"]}


# ── Static frontend ───────────────────────────────────────────────────────────

frontend_dist = os.path.join(os.path.dirname(__file__), "..", "frontend", "dist")
if os.path.exists(frontend_dist):
    # Serve compiled assets directly (JS/CSS/images)
    assets_dir = os.path.join(frontend_dist, "assets")
    if os.path.exists(assets_dir):
        app.mount("/assets", StaticFiles(directory=assets_dir), name="assets")

    # Serve the compiled SPA under an explicit, stable frontend URL as well as
    # the root navigation page. StaticFiles handles /frontend/assets/* safely.
    app.mount("/frontend", StaticFiles(directory=frontend_dist, html=True), name="frontend")

    # SPA fallback for other non-API paths.
    _spa_index = os.path.join(frontend_dist, "index.html")

    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_spa(full_path: str):
        # Exclude API and WS paths from SPA catch-all
        if full_path.startswith("api/") or full_path.startswith("ws/") or full_path.startswith("frontend/"):
            raise HTTPException(status_code=404, detail="Not found")
        candidate = os.path.join(frontend_dist, full_path)
        if os.path.isfile(candidate):
            return FileResponse(candidate)
        return FileResponse(_spa_index)


async def _periodic_director_dispatch():
    """Periodically check and dispatch pending director pipeline groups.
    
    Dispatches groups with classic_status=2 and director_status IN (0, -1, -3)
    that haven't been recently attempted. Retries failed groups (-1, -3) and
    dispatches never-attempted groups (0).
    Skips groups whose SRT files are missing from disk.
    Skips groups that failed with 'no SRT content available' (SRT files are gone).
    """
    from gpu_state import is_online as gpu_is_online
    from transcribe import _run_director_pipeline, _extract_srt_for_director
    BATCH_LIMIT = 10
    while True:
        try:
            if not gpu_is_online():
                logger.debug("Periodic director dispatch: GPU offline — skipping")
                await asyncio.sleep(300)
                continue
            async with aio_connect() as db:
                db.row_factory = aiosqlite.Row
                # Get pending director groups, excluding those with unrecoverable errors
                async with db.execute(
                    """SELECT id, director_error FROM clip_groups
                       WHERE editing_mode = 'director'
                         AND classic_status = 2
                         AND director_status IN (0, -1, -3)
                         AND (director_error IS NULL OR director_error = '' OR (
                           director_error NOT LIKE '%no recordings%'
                           AND director_error NOT LIKE '%recording files missing%'
                           AND director_error NOT LIKE '%physically deleted%'
                           AND director_error NOT LIKE '%无录像文件%'
                           AND director_error NOT LIKE '%无合并素材%'
                           AND director_error NOT LIKE '%no SRT content%'
                           AND director_error NOT LIKE '%SRT%'
                           AND director_error NOT LIKE '%时长%'
                           AND director_error NOT LIKE '%video_clips is EMPTY%'
                         ))
                       ORDER BY id ASC
                       LIMIT ?""",
                    (BATCH_LIMIT,),
                ) as cur:
                    rows = await cur.fetchall()
                    pending_director = [r["id"] for r in rows]

                srt_ready_director = []
                skipped_no_srt = 0
                for gid in pending_director:
                    if await _extract_srt_for_director(gid):
                        srt_ready_director.append(gid)
                    else:
                        await db.execute(
                            "UPDATE clip_groups SET director_status = -2, director_error = ? WHERE id = ?",
                            ("no SRT content available", gid),
                        )
                        skipped_no_srt += 1
                if skipped_no_srt:
                    await db.commit()
                    logger.warning(f"Periodic director dispatch: skipped {skipped_no_srt} groups with missing SRT")
                pending_director = srt_ready_director

                if not pending_director:
                    logger.debug("Periodic director dispatch: no recoverable pending groups")
                else:
                    logger.info(f"Periodic director dispatch: found {len(pending_director)} pending groups")
                    for gid in pending_director:
                        asyncio.create_task(_run_director_pipeline(gid))
                        await asyncio.sleep(0.1)
        except Exception as e:
            logger.warning(f"Periodic director dispatch failed: {e}")
        
        await asyncio.sleep(300)  # Check every 5 minutes


async def _periodic_creative_dispatch():
    """Periodically check and dispatch pending creative pipeline groups.
    
    Only dispatches when GPU service is online. When GPU is offline, skips
    and retries on next cycle.
    """
    from gpu_state import is_online as gpu_is_online
    while True:
        try:
            if not gpu_is_online():
                logger.debug("Periodic creative dispatch: GPU offline — skipping")
                await asyncio.sleep(300)
                continue
            async with aio_connect() as db:
                db.row_factory = aiosqlite.Row
                # Get pending creative groups (status=0) that haven't been dispatched recently
                # Exclude groups with unrecoverable errors
                async with db.execute(
                    """SELECT id FROM clip_groups
                       WHERE classic_status = 2
                         AND director_status = 2
                         AND creative_status = 0
                         AND merged_filename IS NOT NULL
                         AND (creative_error IS NULL OR creative_error = '')
                         AND creative_error NOT LIKE '%duration 0%'
                         AND creative_error NOT LIKE '%no recordings%'
                       ORDER BY id ASC
                       LIMIT 10"""
                ) as cur:
                    pending_creative = [r["id"] for r in await cur.fetchall()]
                
                if pending_creative:
                    logger.info(f"Periodic creative dispatch: found {len(pending_creative)} pending groups")
                    from transcribe import _run_creative_pipeline
                    for gid in pending_creative:
                        asyncio.create_task(_run_creative_pipeline(gid))
                        await asyncio.sleep(0.1)
        except Exception as e:
            logger.warning(f"Periodic creative dispatch failed: {e}")
        
        await asyncio.sleep(300)  # Check every 5 minutes


async def _periodic_qianchuan_dispatch():
    """Periodically check and dispatch pending Qianchuan pipeline groups.

    Dispatches groups with classic_status=2 and qianchuan_status IN (0, -1, -3, -4)
    that haven't been recently attempted. Retries recoverable failures (-1, -3, -4)
    and dispatches never-attempted groups (0).

    Only dispatches when GPU service is online.
    """
    from gpu_state import is_online as gpu_is_online
    BATCH_LIMIT = 5
    while True:
        try:
            if not gpu_is_online():
                logger.debug("Periodic qianchuan dispatch: GPU offline — skipping")
                await asyncio.sleep(300)
                continue
            async with aio_connect() as db:
                db.row_factory = aiosqlite.Row
                # Get pending qianchuan groups, excluding unrecoverable errors
                async with db.execute(
                    """SELECT id FROM clip_groups
                       WHERE classic_status = 2
                         AND qianchuan_status IN (0, -1, -3, -4)
                         AND (qianchuan_error IS NULL OR qianchuan_error = '' OR (
                           qianchuan_error NOT LIKE '%no recordings%'
                           AND qianchuan_error NOT LIKE '%recording files missing%'
                           AND qianchuan_error NOT LIKE '%physically deleted%'
                           AND qianchuan_error NOT LIKE '%无录像文件%'
                           AND qianchuan_error NOT LIKE '%无素材%'
                           AND qianchuan_error NOT LIKE '%时长%'
                           AND qianchuan_error NOT LIKE '%video_clips is EMPTY%'
                         ))
                       ORDER BY id ASC
                       LIMIT ?""",
                    (BATCH_LIMIT,),
                ) as cur:
                    pending_qianchuan = [r["id"] for r in await cur.fetchall()]

                if not pending_qianchuan:
                    logger.debug("Periodic qianchuan dispatch: no recoverable pending groups")
                else:
                    logger.info("Periodic qianchuan dispatch: found %d pending groups", len(pending_qianchuan))
                    # Lazy-import the qianchuan pipeline runner
                    try:
                        from api_v2 import _run_qianchuan_pipeline
                    except ImportError as ie:
                        logger.warning("Periodic qianchuan dispatch: api_v2 not available — skipping (error: %s)", ie)
                        await asyncio.sleep(300)
                        continue
                    for gid in pending_qianchuan:
                        asyncio.create_task(_run_qianchuan_pipeline(gid))
                        await asyncio.sleep(0.1)
        except Exception as e:
            logger.warning("Periodic qianchuan dispatch failed: %s", e)

        await asyncio.sleep(180)  # Check every 3 minutes


async def _periodic_clip_dispatch():
    """Periodically check and dispatch pending clip jobs.
    
    Dispatches recordings with transcribed=2 and clipped=0 that have valid files.
    Only dispatches when GPU service is online.
    """
    from transcribe import _run_editor, get_clip_queue, _pending_heap
    BATCH_LIMIT = 10
    while True:
        try:
            # Check current queue size
            queue_info = get_clip_queue()
            queued_count = len(queue_info.get('queued', []))
            running_count = len(queue_info.get('running', []))
            
            # If queue has many jobs but nothing running, we need to dispatch
            if queued_count > 0 and running_count == 0:
                logger.info(f"Clip dispatch: {queued_count} queued, {running_count} running - checking for new jobs")
            
            if queued_count < 5:  # Only check if queue is nearly empty
                async with aio_connect() as db:
                    db.row_factory = aiosqlite.Row
                    # Get recordings ready for clipping: transcribed=2, clipped=0, not deleted
                    async with db.execute(
                        """SELECT id, filename, clip_count, room_id FROM recordings 
                           WHERE transcribed = 2 AND clipped = 0 
                           AND local_deleted = 0 AND synced = 1
                           ORDER BY end_time ASC
                           LIMIT ?""",
                        (BATCH_LIMIT,),
                    ) as cur:
                        pending_clips = list(await cur.fetchall())
                
                if pending_clips:
                    logger.info(f"Periodic clip dispatch: found {len(pending_clips)} pending clips")
                    for rec in pending_clips:
                        rec_id = rec["id"]
                        filename = rec["filename"]
                        if not filename:
                            continue
                        
                        # Build paths
                        mp4_path = os.path.join(RECORDINGS_DIR, filename)
                        srt_path = os.path.join(RECORDINGS_DIR, 
                            os.path.splitext(filename)[0] + ".srt")
                        
                        # Check files exist
                        if not os.path.exists(mp4_path):
                            logger.debug(f"Clip dispatch: skipping {rec_id}, MP4 missing")
                            continue
                        if not os.path.exists(srt_path):
                            logger.debug(f"Clip dispatch: skipping {rec_id}, SRT missing")
                            continue
                        
                        clip_count = rec["clip_count"] if rec["clip_count"] else 1
                        logger.info(f"Clip dispatch: queuing recording {rec_id} ({filename})")
                        
                        # Create broadcast function for this job
                        async def broadcast(msg):
                            pass  # Silent for background dispatch
                        
                        # Enqueue the job
                        await _run_editor(
                            rec_id, mp4_path, srt_path,
                            clip_count=clip_count,
                            broadcast_fn=broadcast
                        )
                        await asyncio.sleep(0.5)  # Small delay between dispatches
                else:
                    logger.debug("Periodic clip dispatch: no pending clips found")
        except Exception as e:
            logger.warning(f"Periodic clip dispatch failed: {e}")
        
        await asyncio.sleep(60)  # Check every minute


# Add this to the main function
async def main():
    # ... existing code ...
    
    # Start periodic creative dispatch
    cleanup_task = asyncio.create_task(_periodic_cleanup())
    creative_dispatch_task = asyncio.create_task(_periodic_creative_dispatch())
    director_dispatch_task = asyncio.create_task(_periodic_director_dispatch())
    
    # ... rest of existing code ...
    
    cleanup_task.cancel()
    creative_dispatch_task.cancel()
    director_dispatch_task.cancel()
    director_dispatch_task.cancel()
